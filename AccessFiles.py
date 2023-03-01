#pip install openpyxl
import os
import openpyxl
#read file list
#i is the index
def folderfiles(fpath, cd, i):
    #print(fpath)
    #list stored
    res = []
    c = 0
    for item in os.listdir(fpath):
        if os.path.isfile(os.path.join(fpath, item)):
            c += 1
            res.append(item)

    if cd == 'c':
        return c
    elif cd == 'd':
        return res[i]

#fname with extension
#hname is the header name it will locate it position(max 30 rows)
#h is for header row, c is for header column, l is for last row
def excelhloc(fpath, fname, shtn, hname, hlcs, rw):
    fpath = fpath + fname
    wb = openpyxl.load_workbook(fpath)
    sht = wb.worksheets[shtn].title
    fsht = wb.get_sheet_by_name(sht)

    lc = wb[sht].max_column
    #print(hname)
    if hlcs == 'h':
        # provides row of the header at max 30 row search
        for c in range(lc):
            for r in range(31):
                if str(fsht.cell(row=r + 1, column=c + 1).value).strip() == str(hname):
                    ir = r + 1
                    return ir
                elif str(fsht.cell(row=r + 1, column=c + 1).value).strip() != 'None':
                    ir = r + 1
    elif hlcs == 'c' and rw != 0:
        #provides column of the header
        #if no items found default to next column number
        for c in range(lc):
            if str(fsht.cell(row=rw, column=c+1).value).strip() == str(hname):
                ic = c + 1
                return ic
            elif str(fsht.cell(row=rw + 1, column=c + 1).value).strip() != 'None':
                ic = c + 1
                return ic
    elif hlcs == 'l':
        lr = wb[sht].max_row
        return lr
    elif hlcs == 's':
        return sht

#get the data
def excelitem(fpath, fname, sht, r, c):
    fpath = fpath + fname
    wb = openpyxl.load_workbook(fpath)
    fsht = wb.get_sheet_by_name(sht)
    return fsht.cell(row=r, column=c).value
