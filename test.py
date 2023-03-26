import os
import os.path
import shutil
import sys
import time
import zipfile
import builtins
from pathlib import Path
import openpyxl
import numpy as np
import pandas as pd
import pwinput
import pyautogui
import pywinauto
import win32com.client as win32
# import win32com.client as win32com
import xlwings as xw
import xlrd
import io

from msedge.selenium_tools import Edge, EdgeOptions
from openpyxl import load_workbook
from pywinauto import application
from pyxlsb import open_workbook
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait


# global gLatest_file
# global gLatest_FileName
# global gMainWbk
# global gEVFile
# global gProceed
# global dataframe
# global gCurrBL
# global gLocalForUploadFolder
# global gCompleteFolder
# global gFCFolder
# global gCurOnProcBL
# gCurruser=""
# global gDocumentType
# global gMyPAURL
# global gEdgeDriver
# WorkList=[]
# VendorBkg=[]
# UniqueBL=[]

def globalInitializer():
    builtins.gCurruser = getpass.getuser()
    builtins.gLocalForUploadFolder = "C:\\Users\\" + builtins.gCurruser + "\\ADIDocUpload\\"
    #builtins.gFCFolder = "C:\\Users\\" + builtins.gCurruser + "\\DSV\\Adidas Team - Adidas Doc Upload\\For Consolidation"
    builtins.gFCFolder = "C:\\Users\\" + builtins.gCurruser + "\\DSV\\Adidas Team - From L drive\\3. Doc & Scan Tower\\9. Docs Receive\\Automation --DO NOT DELETE --\\Adidas Doc Upload\\For Consolidation"
    builtins.gDownloadFolder = "C:\\Users\\" + builtins.gCurruser + "\\Downloads"
    builtins.gDocumentType = "COMMERCIAL DOCUMENTS"
    #builtins.gCompleteFolder = "C:\\Users\\" + builtins.gCurruser + "\\DSV\\Adidas Team - Adidas Doc Upload\\Uploaded Docs"
    builtins.gCompleteFolder = "C:\\Users\\" + builtins.gCurruser + "\\DSV\\Adidas Team - From L drive\\3. Doc & Scan Tower\\9. Docs Receive\\Automation --DO NOT DELETE --\\Adidas Doc Upload\\Uploaded Docs"
    builtins.gMyPAURL = "https://panalpina.log-net.com/"
    # This serves as the container of Environment variables used across the process. This is where py and xl tool talks with each other
    builtins.gEVFile = "C:\\Users\\" + builtins.gCurruser + "\\ADIDocUpload\\EV.txt"
    builtins.gBookingFile = "C:\\Users\\" + builtins.gCurruser + "\\ADIDocUpload\\BookingList.xlsx"
    builtins.gPQSourceFile = "C:\\Users\\" + builtins.gCurruser + "\\ADIDocUpload\\PQSource.xlsx"
    builtins.gExportedFile = ""
    # This is the MainWbk which serves as UI of the tool. This also does some processes in excel
    builtins.gMainWbk = "C:\\Users\\" + builtins.gCurruser + "\\ADIDocUpload\\DocUploadTool 003.003.5.xlsb"
    builtins.g2bProcXl = "C:\\Users\\" + builtins.gCurruser + "\\ADIDocUpload\\ToBeProcXL.xlsx"
    builtins.gMainWbkName = "DocUploadTool 003.003.5.xlsb"
    builtins.gEdgeDriver = "C:\\Users\\" + builtins.gCurruser + "\\ADIDocUpload\\msedgedriver.exe"
    builtins.VendorBkg = []
    builtins.SOCount = 0
    builtins.gSOGrpType = ""
##

def IsPageReady(edriver):
    PageReady = False
    while PageReady == False:
        try:
            WebDriverWait(edriver, 5).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, '#quickSearchText'))).click()
            return True

        except Exception as e:
            time.sleep(2)
def page_is_loading(driver):
    while True:
        x = driver.execute_script("return document.readyState")
        if x == "complete":
            return True
        else:
            yield False


def IsDirExists(fPath):
    isFExists = False
    if not os.path.exists(fPath):
        isFExists = False
    else:
        isFExists = True
    return isFExists


def makeFolder(FolderPath):
    if not IsDirExists(FolderPath):
        os.makedirs(FolderPath)


def moveFile(srcPath, DestPath):
    if not IsDirExists(DestPath):
        os.rename(srcPath, DestPath)


def deleteFile(FilePath):
    os.remove(FilePath)

def deleteDir(FilePath):
    shutil.rmtree(FilePath)

def ManageCompletedFile(srcPath):
    CompletedFile = ""
    CompletedFPath = ""
    CompletedFile = str(gCurrBL) + ".zip"
    CompletedFPath =  builtins.gCompleteFolder + "\\" + CompletedFile
    if not (IsDirExists(CompletedFPath)):
        shutil.move(srcPath, CompletedFPath)

def GetExportedFile():
    list_of_files = glob.glob(builtins.gDownloadFolder + '/*.xlsx')  # * means all if need specific format then *.csv
    latest_file = max(list_of_files, key=os.path.getctime)
    return latest_file
def FolderZipper(FolderPathToBeZipped, ZippedDest, ZippedName):
    ZippedSuccess=False
    zipobj = zipfile.ZipFile(ZippedDest, 'w', zipfile.ZIP_DEFLATED)
    rootlen = len(FolderPathToBeZipped) + 1
    for base, directories, files in os.walk(FolderPathToBeZipped):
        for file in files:
            fn = os.path.join(base, file)
            zipobj.write(fn, fn[rootlen:])

    if IsDirExists(ZippedDest):
        ZippedSuccess = True
    zipobj.close()
    return ZippedSuccess

# zipfolder(ZippedName, FolderPathToBeZipped)
# sys.exit()
def BtnOKTrials(attemtptCT, eDriver):
    done = False
    if attemtptCT == 0:
        try:
            WebDriverWait(eDriver, 15).until(
                EC.element_to_be_clickable((By.XPATH, '/html/body/div[6]/div[3]/div/button'))).click()
            done = True
            return done
        except TimeoutException:
            return done
    if attemtptCT == 1:
        try:
            WebDriverWait(eDriver, 15).until(
                EC.element_to_be_clickable((By.XPATH, '/html/body/div[5]/div[11]/div/button'))).click()
            done = True
            return done
        except TimeoutException:
            return done
    if attemtptCT == 2:
        try:
            WebDriverWait(eDriver, 15).until(
                EC.element_to_be_clickable((By.XPATH, '/html/body/div[5]/div[3]/div/button'))).click()
            done = True
            return done
        except TimeoutException:
            return done
def BtnClicker(BtnType,eDriver):
    done = False
    BtnTxt = ""
    #Button OK in Add Containers tab
    if BtnType == 'AddctrOK':
        done = OkBtnClicker(eDriver)
        if not done:
            done = BtnOKTrials(0, eDriver)
            if not done:
                done = BtnOKTrials(1, eDriver)
                if not done:
                    done = BtnOKTrials(2, eDriver)
    return done
def OkBtnClicker(eDriver):
    done = False
    try:
        OK_buttons = WebDriverWait(eDriver, 15).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR,
                                                 'body > div.ui-dialog.ui-corner-all.ui-widget.ui-widget-content.ui-front.ui-dialog-buttons.ui-draggable.ui-resizable > div.ui-dialog-buttonpane.ui-widget-content.ui-helper-clearfix > div > button')))
        # check how many buttons in on the HTML, you can try "visibility_of_all_elements_located"
        print(len(OK_buttons))

        visible_buttons = [OK_button for OK_button in OK_buttons if OK_button.is_displayed()]
        visible_buttons_len = len(visible_buttons)

        visible_buttons[visible_buttons_len - 1].click()
        done = True
        return done
    except TimeoutException:
        return done
def ManageOpenDialog(FileToUpload):
    pwa_app = pywinauto.application.Application().connect(title_re=u'Open')
    w_handle = pywinauto.findwindows.find_windows(title=u'Open', class_name='#32770')[0]
    window = pwa_app.window(handle=w_handle)
    ctrl = window['Edit']
    ctrl.set_edit_text(FileToUpload)
    # ctrl.Set_edit_text(FileToUpload)
    # ctrl.SetText(FileToUpload)
    time.sleep(4)
    ctrl2 = window['Open']
    # ctrl2.click_input()
    pwa_app.window(best_match='Open', top_level_only=True).child_window(best_match='O&pen').click()
    time.sleep(3)
    try:
        pwa_app.window(best_match='Open', top_level_only=True).child_window(best_match='Open').click()
        time.sleep(3)
        return True
    except Exception as e:
        return False

    #
    #
    #
    # app = pywinauto.application.Application().connect(title_re=".*Open")
    # mainWindow = app['Open'] # main windows' title
    # ctrl=mainWindow['Edit']
    # mainWindow.SetFocus()
    # ctrl.ClickInput()
    # ctrl.TypeKeys(FileToUpload)
    # ctrlBis = mainWindow['Open'] # open file button
    # ctrlBis.ClickInput()


def Mbox(title, text, style):
    return ctypes.windll.user32.MessageBoxW(0, text, title, style)


def FileZipper(FolderPathToBeZipped, ZippedDest):
    BLzip = zipfile.ZipFile(ZippedDest, 'w')
    for folder, subfolders, files in os.walk(FolderPathToBeZipped):
        for file in files:
            BLzip.write(os.path.join(folder, file), os.path.relpath(os.path.join(folder, file), ZippedDest),
                        compress_type=zipfile.ZIP_DEFLATED)
    BLzip.close()


def zipIt(FolderPathToBeZipped, ZippedDest):
    for root, dirs, files in os.walk(FolderPathToBeZipped):
        for file in files:
            ziph.write(os.path.join(root, file),
                       os.path.relpath(os.path.join(root, file),
                                       os.path.join(FolderPathToBeZipped, '..')))

    with zipfile.ZipFile('Python.zip', 'w', zipfile.ZIP_DEFLATED) as zipf:
        zipIt(ZippedDest + '/', zipf)


def zipToUploadFolder(zippedFilename, zipType, FolderPathToBeZipped, ZippedPath):
    name = zippedFilename
    zip_name = name + '.zip'

    with zipfile.ZipFile(zip_name, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
        for folder_name, subfolders, filenames in os.walk(name):
            for filename in filenames:
                file_path = os.path.join(folder_name, filename)
                zip_ref.write(file_path, arcname=os.path.relpath(file_path, name))

    zip_ref.close()


def unzipIt(zipFilePath, unzippedFilesDest):
    with zipfile.ZipFile(zipFilePath, 'r') as zip_ref:
        zip_ref.extractall(unzippedFilesDest)


def getBkgDetails(eDriver, SO):
    # driver.switch_to.frame('viewFrame')
    DR = ""
    DS = ""
    DEST = ""
    BkgDetails = []
    clName = ""
    WebDriverWait(eDriver, 5).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, '#sections-0Layer > div > div.grid-header > label'))).click()
    SOList = eDriver.find_element(By.ID, "bookingList")
    colCtr = 0
    soValCtr = 0
    soValsub = 0
    soValSub2 = 2
    soValdiv = 2
    elemTxt = ""
    newSO = ""
    elemDRCol = "Doc Receipt Date"
    elemDSCol = "Doc Scan Date"
    elemDestCol = "Destination"
    for Hdrs in eDriver.find_elements(By.CLASS_NAME, "slick-column-name"):
        # find_elements_by_class_name("slick-column-name"):
        colCtr += 1
        elemTxt = Hdrs.text
        elemTxt = elemTxt.strip()

        if (elemDRCol == elemTxt):
            # SOListHdrs.append(hdrs)
            hdrdcol = SOList.text.split("\n")
            for xl in hdrdcol:
                soValCtr += 1
                newSO = xl
                if (soValCtr == colCtr):
                    # builtins.dataframe['DOC RECEIPT DATE'] = newSO
                    soValdiv = colCtr
                    # clName="slick-cell l"+str(soValdiv) +" r" + str(soValdiv) + " lni-editable-cell"
                    clName = "/html/body/div[1]/div[2]/div[2]/div[1]/div/div[2]/div[2]/div[2]/div[2]/div/div/div/div[2]/div[5]/div/div[1]/div[" + str(
                        soValdiv) + "]"
                    DR = GetFirstInputElemByXpathInTbl(clName, 'DOC RECEIPT DATE', eDriver)
                    # builtins.dataframe['DOC RECEIPT DATE'] = GetFirstInputElemByXpathInTbl(clName, 'DOC RECEIPT DATE',eDriver)

                    break
            # BkgDetails.append(newSO)
        if (elemDSCol == elemTxt):
            # SOListHdrs.append(hdrs)
            hdrdcol = SOList.text.split("\n")
            for xl in hdrdcol:
                soValCtr += 1
                newSO = xl
                if (soValCtr == colCtr):
                    clName = "/html/body/div[1]/div[2]/div[2]/div[1]/div/div[2]/div[2]/div[2]/div[2]/div/div/div/div[2]/div[5]/div/div[1]/div[" + str(
                        colCtr) + "]"
                    DS = GetFirstInputElemByXpathInTbl(clName, 'DOC SCAN DATE', eDriver)
                    # builtins.dataframe['DOC SCAN DATE'] = GetFirstInputElemByXpathInTbl(clName, 'DOC SCAN DATE', eDriver)
                    break
            # BkgDetails.append(newSO)
        if (elemDestCol == elemTxt):
            clName = "/html/body/div[1]/div[2]/div[2]/div[1]/div/div[2]/div[2]/div[2]/div[2]/div/div/div/div[2]/div[5]/div/div[1]/div[" + str(
                colCtr) + "]"
            DEST = GetFirstInputElemByXpathInTbl(clName, 'DEST', eDriver)
            # builtins.dataframe['DEST'] = GetFirstInputElemByXpathInTbl(clName, 'DEST', eDriver)

    #update DR, DS & DEST to DF
    # soIndx = 0
    # for row in builtins.dataframe.itertuples():
    #     if row.SO == SO:
    #         l = row.Index
    #         soIndx = l
    #         builtins.dataframe.loc[soIndx, 'DOC RECEIPT DATE'] = "."
    #         builtins.dataframe.loc[soIndx, 'DOC SCAN DATE'] = "."
    #         builtins.dataframe.loc[soIndx, 'DEST'] = "."

    updateBkgDetails(eDriver, SO, DR, DS, DEST)

    # for row in builtins.dataframe.itertuples():
    #     if row.SO == SO:
    #         builtins.dataframe.loc[soIndx, 'DOC RECEIPT DATE'] = DR
    #         builtins.dataframe.loc[soIndx, 'DOC SCAN DATE'] = DS
    #         builtins.dataframe.loc[soIndx, 'DEST'] = DEST
    #         soIndx += 1
    # return BkgDetails
def updateBkgDetails(edriver, SO, DR, DS, DEST):
    soIndx = -1
    l = 0
    if DR=="":
        DR="."
    if DS=="":
        DS="."
    # for i, row in enumerate(builtins.dataframe.values):
    #     if builtins.dataframe.loc[i, 'SO'] == SO:
    #         builtins.dataframe.loc[i, 'DOC_RECEIPT_DATE'] = DR
    #         builtins.dataframe.loc[i, 'DOC_SCAN_DATE'] = DS
    #         builtins.dataframe.loc[i, 'DEST'] = DEST

    for row in builtins.dataframe.itertuples():
        if row.SO == SO:
            l = row.Index
            soIndx = l
            builtins.dataframe.loc[soIndx, 'DOC_RECEIPT_DATE'] = DR
            builtins.dataframe.loc[soIndx, 'DOC_SCAN_DATE'] = DS
            builtins.dataframe.loc[soIndx, 'DEST'] = DEST

    # for row in builtins.dataframe.itertuples():
    #     soIndx += 1
    #     if row.SO == SO:
    #         builtins.dataframe.loc[soIndx, 'DOC RECEIPT DATE'] = DR
    #         builtins.dataframe.loc[soIndx, 'DOC SCAN DATE'] = DS
    #         builtins.dataframe.loc[soIndx, 'DEST'] = DEST

def findVendorBkgText(edriver, className):
    vdrBkg = []

    elemTst = "Vendor Booking"
    colCtr = 0
    hdrColName = ""
    hdrColTxt = ""

    for Hdrs in edriver.find_elements(By.CLASS_NAME, "grid-canvas"):
        colCtr += 1
        hdrColTxt = Hdrs.text
        hdrColName = hdrColTxt.strip()
        if hdrColName == elemTst:
            return colCtr

    for SOtext in edriver.find_elements(By.CLASS_NAME, "slick-cell l8 r8"):
        newSOC = ""
        newSOC = SOtext.strip()
        if not (newSOC in vdrBkg):
            vdrBkg.append(newSOC)

    return vdrBkg
def findVendorBkgCol(edriver):
    elemTst = "Vendor Booking"
    colCtr = 0
    hdrColName = ""
    hdrColTxt = ""
    for Hdrs in edriver.find_elements(By.CLASS_NAME, "slick-column-name"):
        colCtr += 1
        hdrColTxt = Hdrs.text
        hdrColName = hdrColTxt.strip()
        if hdrColName == elemTst:
            return colCtr
def getVdrBkgMultiple(edriver):
    # edriver.switch_to.default_content()
    # edriver.switch_to.frame('viewFrame')

    vdrBkg = []
    colindx = 0

    # get vendorbooking column index
    colindx = findVendorBkgCol(edriver)
    VdrBkgColClass = ""
    elem = ""
    newSO = ""
    colSub = 1
    gridVBKgColNdx = colindx-colSub
    lpref=""
    rpref=""
    lpref = "slick-cell l" + str(gridVBKgColNdx)
    rpref = "r" + str(gridVBKgColNdx)
    VdrBkgColClass = lpref + ' '
    VdrBkgColClass = VdrBkgColClass + rpref
    # VdrBkgColClass = "slick-cell l" + str(gridVBKgColNdx) + " " + "r" + str(gridVBKgColNdx)

    # for elem in edriver.find_elements(By.CLASS_NAME, VdrBkgColClass):
    solist = []

    VdrBkgList = edriver.find_elements(By.CLASS_NAME, "slick-cell l8 r8")
    for newSOA in VdrBkgList:
        newSOC = ""
        newSOC = newSOA.strip()
        if not (newSOC in vdrBkg):
            vdrBkg.append(newSOC)

    vdrBkg = findVendorBkgText(edriver, VdrBkgColClass)
    return vdrBkg


def getVdrBkg(edriver):
    # driver.switch_to.frame('viewFrame')
    vdrBkg = []


    colCtr = 0
    soValCtr = 0
    soValsub = 0
    soValSub2 = 2
    soValdiv = 1
    elemTxt = ""
    newSO = ""
    elemTst = "Vendor Booking"

    if(IsStrBySelectorVisibleOnUI('BkgProperties', edriver)):
        # vdrBkg = GetInputElemValByID('shpOptionalValHiddenField1', edriver)
        vdrBkg =GetInputElemValByNAME('shpOptionalValHiddenField1', edriver)
        vdrBkg = edriver.find_element(By.CSS_SELECTOR('#shpOptionalValHiddenField1'))
        vdrBkg = vdrBkg.text
    else:
        SOList = edriver.find_element(By.CLASS_NAME, "grid-canvas")
        for Hdrs in edriver.find_elements(By.CLASS_NAME, "slick-column-name"):
            # find_elements_by_class_name("slick-column-name"):
            # colCtr += 1

            elemTxt = Hdrs.text
            elemTxt = elemTxt.strip()

            # if (elemTst == elemTxt):
                # SOListHdrs.append(hdrs)
            hdrdcol = SOList.text.split("\n")

            if(colCtr>len(hdrdcol)):
                return vdrBkg
            else:
                soSrcTxt = 'Open'
                soValsub = colCtr - soValdiv
                for indx, elem in enumerate(hdrdcol):
                    colCtr += 1
                    soValCtr = 0
                    # soValCtr = soValCtr + 1
                    if elem == soSrcTxt:
                        soValCtr = indx + soValSub2
                        newSO = hdrdcol[soValCtr]
                        # newSO = xl
                        # if (soValCtr == soValsub):
                        if not (newSO in vdrBkg):
                            soValsub = soValsub + 18
                            vdrBkg.append(newSO)
                    if (elem=='Delete'):
                        colCtr = 0

    return vdrBkg

def GetMultipleSO(eDriver):
    downloadSuccess = False
    dwnLdWaitCtr = 0
    vdrBkg = []
    initExpFile = ""
    expFile = ""
    #get the initial content of Downloaded folder before trigerring another donload
    initExpFile = GetExportedFile()
    eDriver.execute_script('parent.lognetScreen.resultList.exportRows()')
    time.sleep(5)
    while not downloadSuccess:
        expFile = GetExportedFile()
        dwnLdWaitCtr += 1
        if expFile != initExpFile:
            downloadSuccess = True
            break
        else:
            if dwnLdWaitCtr >3 :
                expFile = ""
                break
    builtins.gExportedFile = expFile
    # builtins.exportedDataframe = pd.read_excel(expFile, 'Search Export')
    # builtins.exportedDataframe = openpyxl.load_workbook(expFile)
    # builtins.exportedDataframe = pd.read_excel(expFile, engine='openpyxl')
    # make the PQSource file which will be the data source of Power Query table
    if(IsDirExists(builtins.gPQSourceFile)):
        deleteFile(builtins.gPQSourceFile)
    if not expFile=="":
        moveFile(expFile, builtins.gPQSourceFile)

    else:
        Mbox('ADI Doc Upload Tool v.7.2', 'Network Problem while accessing MyPA, Pls try again later', 0)
        return vdrBkg
    # refresh power query
    # xlapp = win32com.gencache.EnsureDispatch('Excel.Application')
    xlapp = win32.DispatchEx('Excel.Application')
    wb = xlapp.Workbooks.Open(builtins.gBookingFile)
    wb.RefreshAll()
    # xlapp.CalculateUntilAsyncQueriesDone()

    time.sleep(7)
    wb.Save()
    # wb.Close()
    xlapp.Quit()
    time.sleep(4)
    builtins.exportedDataframe = pd.read_excel(open(builtins.gBookingFile, 'rb'),
                                               sheet_name='Search Export', usecols="A", header=0)
    # worksheet = workbook.sheet_by_index(0)

    vdrBkg = builtins.exportedDataframe  # .F.unique
    builtins.SOCount = len(vdrBkg)
    return vdrBkg
# this is to set Enviromental variable "DNLDFILE" to be used by MainWBK --> the Main excel macro
def UpdateEV_DF(DFPath, DFname=""):
    # mainWbk=openpyxl.load_workbook(wbkPath, read_only=false, keep_vba=true)
    # mainWS = srcfile.get_sheet_by_name('Main')
    # mainWS['P5']=dwnlddFname
    os.environ["DWNLDFILE"] = DFPath
    print(os.environ["DWNLDFILE"])


# this is the function to trigger the function DoTheExcelProcess in MainWbk
def RunnerMainWBK(DFPath, gCurrBL):
    xl = win32com.client.Dispatch("Excel.Application")
    xl.Workbooks.Open(Filename=builtins.gMainWbk, ReadOnly=1)
    xl.Application.Run("DocUploadTool 003.003.5.xlsb!MainProcess.DoTheExcelProcess", DFPath, gCurrBL)
    xl.Application.Quit()  # Comment this out if your excel script closes
    del xl


def MainWBkWriter():
    app = xw.App()
    mainWB = xw.Book(builtins.gMainWbkName)
    procShName = 'Processed'
    procSht = mainWB.sheets[procShName]
    procShtLstRow = mainWB.sheets[procShName].range('A' + str(mainWB.sheets[procShName].cells.last_cell.row)).end(
        'up').row
    procShtLstRow += 1
    ctbUpdated = 'A' + str(procShtLstRow)

    # procSht.range(procShtLstRow).api.Delete(DeleteShiftDirection.xlShiftUp)
    procSht.range("A1", "P" + str(procShtLstRow)).value = ''
    mainWB.sheets[procShName].range('A1').value = builtins.dataframe
    procSht.range('A' + ctbUpdated).value = builtins.dataframe


def MainWbkReader():
    app = xw.App()
    mainWB = xw.Book(builtins.gMainWbkName)
    mainShName = 'Main'
    mainSht = mainWB.sheets[mainShName]
    procShtLstRow = mainWB.sheets[mainShName].range('A' + str(mainWB.sheets[mainShName].cells.last_cell.row)).end(
        'up').row


def ReadEV(EVkey):
    lineNum = 0
    searchRes = ""
    EVkey = EVkey.lower()
    linenum = 0
    with open(builtins.gEVFile, 'rt') as myEVfile:
        for line in myEVfile:
            linenum += 1
            if line.lower().find(EVkey) != -1:
                searchRes = line.rstrip('\n')
                return searchRes


# this is to open External Document tab in MyPA+ to prepare for processes related to this tab
# def ManageOpeningExtDoc():
# driver.switch_to.default_content()
# driver.switch_to.frame('viewFrame')
# click Documentation Menu
# WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="menuBar"]/li[5]/a'))).click()
# time.sleep(2)
# menu = driver.find_element(By.CSS_SELECTOR('#menuBar > li:nth-child(5) > a'))
# hidden_submenu = driver.find_element(By.CSS_SELECTOR('#menuBar > li:nth-child(5) > ul > li:nth-child(9) > ul > li:nth-child(2) > a'))
#
# actions = ActionChains(driver)
# actions.move_to_element(menu)
# actions.click(hidden_submenu)
# actions.perform()
#
# click External Document
# WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="menuBar"]/li[5]/ul/li[9]'))).click()
# time.sleep(1)
# click Find External Document
# WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="menuBar"]/li[5]/ul/li[9]/ul/li[2]/a'))).click()
# ext_menu = driver.find_element(By.XPATH('//*[@id="menuBar"]/li[5]/ul/li[9]/a'))
# ext_menu.click()

# wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '#menuBar > li:nth-child(5) > ul > li:nth-child(9) > a'))).click()
# wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '#menuBar > li:nth-child(5) > ul > li:nth-child(9) > ul > li:nth-child(2) > a'))).click()

# WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div[2]/div[1]/div/ul[3]/li[5]/ul/li[9]/a'))).click()
# ext_menu = driver.find_element(By.XPATH('//*[@id="menuBar"]/li[5]/ul/li[9]/ul/li[2]/a'))
# ext_menu.click()
# WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div[2]/div[1]/div/ul[3]/li[5]/ul/li[9]/ul/li[2]/a'))).click()
# ----------------
# actions = ActionChains(driver)
# ext_menu = driver.find_element(By.XPATH('//*[@id="menuBar"]/li[5]/ul/li[9]/a'))
# actions.move_to_element(ext_menu)
# actions.clickAndHold()
# -----------------
# #WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#menuBar > li:nth-child(5) > ul > li:nth-child(9) > a'))).click()
# # menuBar > li:nth-child(5) > ul > li:nth-child(9) > ul > li:nth-child(2) > a
# WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'menuBar > li:nth-child(5) > ul > li:nth-child(9) > ul > li:nth-child(2) > a'))).click()
# driver.execute_script('top.gotoScreen(\'GetSearch\', \'ExternalDocument\', \'SEARCH\', true, \'ld_document_type=EXTERNAL\')')
# time.sleep(2)
def OpenExternalDocs(eDriver):
    # ManageOpeningExtDoc()
    # click Documentation tab
    eDriver.switch_to.default_content()
    eDriver.switch_to.frame('viewFrame')
    # WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="menuBar"]/li[5]/a'))).click()
    # click External Docs
    # time.sleep(1)
    # driver.find_element(By.XPATH('//*[@id="menuBar"]/li[5]/ul/li[9]/a/span')).click()
    # WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="menuBar"]/li[5]/ul/li[9]/a'))).click()
    # click Find External Docs
    # WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="menuBar"]/li[5]/ul/li[9]/ul/li[2]/a'))).click()
    # time.sleep(1)
    # driver.find_element(By.XPATH('//*[@id="menuBar"]/li[5]/ul/li[9]/ul/li[2]/a')).click()
    # time.sleep(1)

    eDriver.execute_script(
        'top.gotoScreen(\'GetSearch\', \'ExternalDocument\', \'SEARCH\', true, \'ld_document_type=EXTERNAL\')')

    time.sleep(2)


# this is to Open FindWhs Bookings tab in MyPA+
def OpenFindWhsBkg(eDriver):
    eDriver.execute_script("top.gotoScreen('GetSearch', 'Booking', 'SEARCH', true, '')", "JavaScript")
    ##WhsBkg=driver.find_element(By.CLASS_NAME, "sf-with-ul")
    ##a = ActionChains(driver)
    ##m = driver.find_element_by_link_text("Booking")
    ##a.move_to_element(m).perform()
    # WhsBkg.click()
    time.sleep(5)
    # WebDriverWait(eDriver, 4).until(
    #     EC.element_to_be_clickable((By.CSS_SELECTOR, '#quickSearchText'))).click()
    # if (IsPageReady(eDriver)):
    #     WebDriverWait(eDriver, 4).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="menuBar"]/li[2]/a'))).click()
    #    # WebDriverWait(driver, 4).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div[2]/div[1]/div/ul[3]/li[2]/ul/li[2]/a'))).click()


def CreateExternalDocs(BLNum, ZippedFileLocation, UploadType, UploadRemarks,eDriver):
    uploadType = ""
    txtPartial = 'PARTIAL'
    uploadType = UploadType
    IsPartial = False
    if txtPartial in UploadType:
        IsPartial = True
    # open create external docs tab
    # click Documentation tab
    WebDriverWait(eDriver, 2).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="menuBar"]/li[5]/a'))).click()
    # click External Docs
    WebDriverWait(eDriver, 2).until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="menuBar"]/li[5]/ul/li[9]/a'))).click()
    # click Find External Docs
    # WebDriverWait(driver, 1).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="menuBar"]/li[5]/ul/li[9]/ul/li[2]/a'))).click()
    # driver.execute_script('top.gotoScreen(\'GetSearch\', \'ExternalDocument\', \'SEARCH\', true, \'ld_document_type=EXTERNAL\'\)')
    eDriver.execute_script(
        'top.gotoScreen(\'GetEntry\', \'ExternalDocument\', \'ENTRY\', true, \'ld_document_type=EXTERNAL\')')
    time.sleep(5)

    WebDriverWait(eDriver, 6).until(EC.element_to_be_clickable((By.ID, 'docId'))).send_keys(BLNum)

    # select = Select(driver.find_element(By.XPATH('//*[@id="edFormPanelLayer"]/form/table/tbody/tr[8]/td[2]/nobr/select')))
    # For Partial Upload Type Complete Docs field should be No, otherwise it's Yes
    if (IsPartial):
        WebDriverWait(eDriver, 3).until(
            EC.element_to_be_clickable(By.XPATH, '//*[@id="edFormPanelLayer"]/form/table/tbody/tr[8]/td[2]/nobr/select/option[3]')).click()
        # driver.execute_script('parent.setIsModified(\'lognetScreen\', true); parent.gui.screen.dataMap.setValue(\'ldo.opt_field_10683\', this.options[this.selectedIndex].value, 0, null, null);')
        # WebDriverWait(driver, 4).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="docId"]'))).send_keys('No')
        # driver.execute_script('parent.setIsModified(\'lognetScreen\', true); parent.gui.screen.dataMap.setValue(\'ldo.opt_field_10683\', this.options[this.selectedIndex].value, 0, null, null);')
        WebDriverWait(eDriver, 4).until(
            EC.element_to_be_clickable((By.XPATH, 'o//*[@id="optionalOptionalValHiddenField1"]'))).send_keys(
            UploadRemarks)
    #else:
        # WebDriverWait(eDriver, 2).until(EC.element_to_be_clickable(
        #     (By.XPATH, '//*[@id="edFormPanelLayer"]/form/table/tbody/tr[8]/td[2]/nobr/select/option[2]'))).click()
        # # UploadTypeSelect.select_by_visible_text('Yes')
        # # driver.execute_script('parent.setIsModified(\'lognetScreen\', true); parent.gui.screen.dataMap.setValue(\'ldo.opt_field_10683\', this.options[this.selectedIndex].value, 0, null, null);')
    # Remove Missing Doc Details remarks for COMPLETE upload type
    if (uploadType.find("COMPLETE")) > 0:
        WebDriverWait(eDriver, 4).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="optionalOptionalValHiddenField1"]"]'))).clear()
    # Input Document Type
    # documentType = "COMMERCIAL INVOICE"
    WebDriverWait(eDriver, 4).until(EC.element_to_be_clickable((By.ID, 'docDesc'))).send_keys(builtins.gDocumentType)
    # Fillup line Items fields
    WebDriverWait(eDriver, 4).until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="tabSlideMainPanelLayer"]/table/tbody/tr[4]/td/a'))).click()
    eDriver.execute_script('parent.gui.screen.docTab.changeSlide(1);')
    time.sleep(2)
    ManageLineItemsTab(BLNum, eDriver)
    # ManageUploadTab(BLNum)


def ManageLineItemsTab(BLNum, eDriver):
    eDriver.switch_to.default_content()
    eDriver.switch_to.frame('viewFrame')
    # driver.switch_to.frame('viewFrame')
    # content = driver.find_element(By.NAME,"shipmentsFormPanel")
    WebDriverWait(eDriver, 5).until(EC.element_to_be_clickable(
        (By.XPATH, '//*[@id="shipmentsFormPanelLayer"]/form/table/tbody/tr[2]/td[3]/nobr/select/option[6]'))).click()
    ##selItem= driver.find_element(By.NAME('selectionField'))
    ##selectItems=Select(selItem)

    # selectItems=Select(driver.find_element(By.XPATH('//*[@id="shipmentsFormPanelLayer"]/form/table/tbody/tr[2]/td[3]/nobr/select')))
    # selectLineItems = Select(WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/div[2]/div[2]/div[1]/div/div[2]/div[2]/div[2]/div[2]/div/div/div/div/table/tbody/tr/td/div/table/tbody/tr[2]/td/div/table/tbody/tr/td/div/table/tbody/tr[2]/td/div/form/table/tbody/tr[2]/td[3]/nobr/select'))))
    # selectLineItems.select_by_visible_text('Container')
    time.sleep(2)
    # driver.execute_script('parent.setIsModified(\'lognetScreen\', true); parent.gui.screen.dataMap.setValue(\'selectOption\', this.options[this.selectedIndex].value, 0, null, null);parent.gui.screen.selectionField.determineRedraw(); parent.gui.screen.selectionField.forceRedraw();')
    WebDriverWait(eDriver, 3).until(EC.alert_is_present())
    alert = eDriver.switch_to.alert
    alert_text = alert.text
    # alert.accept()

    # if((alert_text).find("Choosing a new value")>0):
    alert.accept()
    time.sleep(2)
    WebDriverWait(eDriver, 2).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="selectedLookup"]'))).click()
    time.sleep(2)
    # Update BL Num to Bill of Lading field
    WebDriverWait(eDriver, 3).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="dynamicField6"]'))).send_keys(BLNum)
    WebDriverWait(eDriver, 5).until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="searchControlCriteriaButtonPanelLayer"]/input'))).click()
    time.sleep(3)
    # click Select All
    eDriver.execute_script('top.gui.screen.rw.selectAllContainers();')
    time.sleep(2)
    eDriver.switch_to.default_content()
    eDriver.switch_to.frame('viewFrame')
    time.sleep(5)
    # click OK button
    doneClick = False
    try:
        if not BtnClicker('AddctrOK', eDriver):
            time.sleep(3)
            try:
                WebDriverWait(eDriver, 10).until(
                                EC.element_to_be_clickable((By.XPATH, '/html/body/div[5]/div[3]/div/button'))).click()
            except TimeoutException:
                try:
                    WebDriverWait(eDriver, 10).until(
                                    EC.element_to_be_clickable((By.XPATH, '/html/body/div[5]/div[11]/div/button'))).click()
                except TimeoutException:
                    try:
                        WebDriverWait(eDriver, 10).until(
                            EC.element_to_be_clickable((By.XPATH, '/html/body/div[6]/div[3]/div/button'))).click()
                        doneClick = True
                    except TimeoutException:
                        try:
                            WebDriverWait(eDriver, 10).until(
                                EC.element_to_be_clickable((By.XPATH, '/html/body/div[6]/div[11]/div/button'))).click()
                            doneClick = True
                        finally:
                            doneClick = False
            except TimeoutException:
                try:
                    WebDriverWait(eDriver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, '/html/body/div[5]/div[11]/div/button'))).click()
                except TimeoutException:
                        OkBtnTxt = 'OK'
                        BtnTxt = ""
                        for Btn in eDriver.find_elements(By.CLASS_NAME, "ui-button ui-corner-all ui-widget"):
                            BtnTxt = Btn.text
                            BtnTxt = BtnTxt.strip()
                            if BtnTxt == OkBtnTxt:
                                Btn.click()
    except TimeoutException:
        EC.element_to_be_clickable(By.CSS_SELECTOR,'body > div.ui-dialog.ui-corner-all.ui-widget.ui-widget-content.ui-front.ui-dialog-buttons.ui-draggable.ui-resizable > div.ui-dialog-buttonpane.ui-widget-content.ui-helper-clearfix > div > button').click()
        # if IsStrByXPathVisibleOnUI('OK',eDriver):
        #     WebDriverWait(eDriver, 5).until(
        #         EC.element_to_be_clickable((By.XPATH, '/html/body/div[5]/div[3]/div/button'))).click()
        # else:
        #     WebDriverWait(eDriver, 5).until(
        #         EC.element_to_be_clickable((By.XPATH, '/html/body/div[5]/div[11]/div/button'))).click()
        time.sleep(4)
        # buttons = driver.find_elements_by_xpath("//*[contains(text(), 'OK')]")
        #
        # for btn in buttons:
        #     btn.click()
        ##WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[5]/div[11]/div/button'))).click()
    #else:
    # click Submit button
    WebDriverWait(eDriver, 2).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="screenToolBar"]/a[1]'))).click()
    time.sleep(4)


# This function is used to handle upload process
def ManageUploadTab(BLNum, FileToUpload,eDriver):
    eDriver.switch_to.default_content()
    eDriver.switch_to.frame('viewFrame')
    soIndx = 0
    # click Files to Upload tab
    WebDriverWait(eDriver, 2).until(EC.element_to_be_clickable(
        (By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[1]/div/div[2]/div[2]/div[2]/div[2]/ul/li[2]/a'))).click()
    time.sleep(2)

    eDriver.switch_to.frame('uploadIFrame')
    # s=driver.find_element(By.XPATH('//input[@type=''File'']'))
    # s.send_keys(FileToUpload)
    # time.sleep(2)
    # WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[1]/div/div[2]/div[2]/div[2]/div[2]/div/div[2]/div[2]/div/form/table/tbody/tr[1]/td[2]/input'))).click()

    # WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.NAME,'fileUploadX1'))).click()
    # WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#fileForm > form > table > tbody > tr:nth-child(1) > td:nth-child(2) > input[type=FILE]"))).click()

    # elements = driver.find_elements(By.XPATH('//*[@id="fileForm"]/form/table/tbody/tr[1]/td[2]/input'))

    # WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="fileForm"]/form/table/tbody/tr[1]/td[2]/input'))).click()
    # elements=WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="fileForm"]/form/table/tbody/tr[1]/td[2]/input')))
    # for element in elements:
    #    driver.execute_script("arguments[0].click();", element);
    #     print('Pass')

    # elements = driver.find_elements(By.CSS_SELECTOR("input[type='FILE']"))
    # filebutt =WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR("input[type='file']"))))
    # filebutt.click()
    # for element in elements:
    #     driver.execute_script("arguments[0].click();", element);
    #     element.click()
    #     break

    # WebDriverWait(self.driver, 10).until(element_present).click()  # This opens the windows file selector
    actions = ActionChains(eDriver)
    nthtry = 1
    for _ in range(nthtry):
        actions = actions.send_keys(Keys.TAB)
        time.sleep(nthtry)
        actions = actions.send_keys(Keys.ENTER)
        actions.perform()
    time.sleep(6)
    # pyautogui.getWindowsWithTitle("Open")[0].
    ManageOpenDialog(FileToUpload)
    time.sleep(1)
    eDriver.switch_to.default_content()
    eDriver.switch_to.frame('viewFrame')
    WebDriverWait(eDriver, 2).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="screenToolBar"]/a[1]'))).click()
    time.sleep(3)
    WebDriverWait(eDriver, 5).until(EC.alert_is_present())
    alert = eDriver.switch_to.alert

    alertTxt = alert.text
    if alertTxt.find('successfully') > 0:
        UpdateToolIFRemarks(gCurrBL,'UPLOADED')
        #builtins.dataframe.loc[soIndx, 'TOOL_REMARKS'] = 'UPLOADED'  # alert.text
        #builtins.dataframe.loc[soIndx, 'STATUS'] = 'UPLOADED'
        alert.accept()
        ManageCompletedFile(FileToUpload)

    # element_present =WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH, '//*[@id="fileForm"]/form/table/tbody/tr[1]/td[2]/input')))
    # WebDriverWait(self.driver, 20).until(element_present).click()  # This opens the windows file selector
    # pyautogui.write(FileToUpload)
    # pyautogui.press('return')

    # WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="documentSearchControlCriteriaButtonPanelLayer"]/input'))).click()

    # don't click on the element, just send the path directly
    # element.send_keys(FileToUpload)
    # time.sleep(2)
    # s = driver.find_element(By.XPATH, "//input[@type='file']")
    ## file path specified with send_keys
    # s.send_keys(FileToUpload)
    # s.send_keys("O")
    # time.sleep(2)
    # WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="screenToolBar"]/a'))).click()


# this function will perform post processing tasks such as filling builtins.dataframe data for PIC and DATE UPLOADED,this will also data dump to Main Wbk and finally show an lert to let the user know that the process has ended
def ManagePostProcessing(BLNum):
    gCurruser = getpass.getuser()
    soIndx = 0
    currDT = datetime.datetime.now()
    for row in builtins.dataframe.itertuples():
        soIndx=0
        if not BLNum == '.':
            if row.BL == BLNum:
                soIndx = row.Index
                builtins.dataframe.loc[soIndx, 'PIC'] = gCurruser
                builtins.dataframe.loc[soIndx, 'DATE_UPLOADED'] = currDT
                if(builtins.dataframe.loc[soIndx, 'TOOL_REMARKS']) == 'UPLOADED':
                    DocTobeUploaded = ""
                    DocTobeUploaded = builtins.gCompleteFolder + "\\" + BLNum + '.zip'
                    #move uploaded file to Teams Uploaded folder
                    if IsDirExists(DocTobeUploaded):
                        UpdateToolIFRemarks(BLNum, "Already Synced")
                    else:
                        try:
                            shutil.move(builtins.dataframe.loc[soIndx, 'UPLOADED_FILE'], builtins.gCompleteFolder)
                        except:
                            UpdateToolIFRemarks(BLNum, "Already Synced")
        else:
            if (builtins.dataframe.loc[soIndx, 'TOOL_REMARKS']) == 'UPLOADED':
                BLNum = builtins.dataframe.loc[soIndx, 'BL']
                builtins.dataframe.loc[soIndx, 'PIC'] = gCurruser
                builtins.dataframe.loc[soIndx, 'DATE_UPLOADED'] = currDT
                DocTobeUploaded = ""
                DocTobeUploaded = builtins.gCompleteFolder + "\\" + BLNum + '.zip'
                # move uploaded file to Teams Uploaded folder
                if IsDirExists(DocTobeUploaded):
                    UpdateToolIFRemarks(BLNum, "Already Synced")
                else:
                    try:
                        shutil.move(builtins.dataframe.loc[soIndx, 'UPLOADED_FILE'], builtins.gCompleteFolder)
                    except:
                        UpdateToolIFRemarks(BLNum, "Already Synced")
                    # moveFile(builtins.dataframe.loc[soIndx, 'UPLOADED_FILE'], builtins.gCompleteFolder)
                    # #delete prev downloaded MyPA file
                    # if IsDirExists(builtins.gExportedFile):
                    #     deleteFile(builtins.gExportedFile)
            soIndx+=1
    MainWBkWriter()

    Mbox('ADI Doc Upload Tool v.7.2', 'Process Completed! Pls see Processed Tab for details.', 0)


# This function is used to perform tasks related to OpenFindWhsBkg
def ManageExternalDocs(BLNum, eDriver):
    soIndx = 0
    SearchResult = ""
    ZippedFileLoc = ""
    UploadType = ""
    UploadRemarks = ""
    for row in builtins.dataframe.itertuples():
        if row.BL == BLNum:
            soIndx = row.Index
            txtConsolidated = "Consolidated"
            txtZipped = "Zipped"
            #if (builtins.dataframe.loc[soIndx, 'TOOL_REMARKS']) == 'Consolidated|Consolidated|Zipped':
            if not CheckRemarksIfComplete(BLNum):
                if txtConsolidated in builtins.dataframe.loc[soIndx, 'TOOL_REMARKS']:
                    if txtZipped in builtins.dataframe.loc[soIndx, 'TOOL_REMARKS']:
                        ZippedFileLoc = builtins.dataframe.loc[soIndx, 'UPLOADED_FILE']
                        UploadType = builtins.dataframe.loc[soIndx, 'STATUS']
                        UploadRemarks = builtins.dataframe.loc[soIndx, 'REMARKS']
                         # if math.isnan(ZippedFileLoc):
                        # ZippedFileLoc=
                        if not (builtins.dataframe.loc[soIndx, 'STATUS']=='COMPLETE UPLOAD'):
                            if UploadRemarks:
                                if not (math.isnan(UploadRemarks)):
                                    rem = UploadRemarks
                                else:
                                    UploadRemarks = 'Incomplete Docs'
                            else:
                                UploadRemarks = 'Incomplete Docs'
                        if IsPageReady(eDriver):
                            eDriver.find_element(By.ID, 'dynamicField1').send_keys(gCurrBL)
                            WebDriverWait(eDriver, 5).until(EC.element_to_be_clickable((By.NAME, 'searchButton'))).click()
                            time.sleep(2)
                            createExtDoc = False

                            createExtDoc = (IsProceedToExtDocCreation(eDriver))
                    else:
                        createExtDoc = False
            else:
                createExtDoc = False
                # if (IsStrByXPathVisibleOnUI('//*[@id="sections-0Layer"]/div[1]')):
                #     createExtDoc = False
                # if (IsStrByXPathVisibleOnUI('//*[@id="searchResultsListShowFilterButton"]/nobr')):
                #     createExtDoc = False
                #
                # if not((IsStrByIDVisibleOnUI("norecords"))):
                #     createExtDoc = True
                #
                # if (IsStrByIDVisibleOnUI("norecords")):
                #     createExtDoc=True
                # #if not(IsStrByIDVisibleOnUI("searchResultsListPanel")):
            if createExtDoc:
                CreateExternalDocs(BLNum, ZippedFileLoc, UploadType, UploadRemarks, eDriver)
                ManageUploadTab(gCurrBL, ZippedFileLoc, eDriver)
                eDriver.switch_to.default_content()
                eDriver.switch_to.frame('viewFrame')
                time.sleep(2)
                ##WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[1]/div/div[2]/div[2]/div[1]/div[2]/a[1]'))).click()
                # WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="screenToolBar"]/a'))).click()
                # time.sleep(2)
                # driver.execute_script("""top.executeSlickGridFunctions(); parent.gui.screen.submit()""")
                # time.sleep(3)
                # WebDriverWait(driver, 3).until(EC.alert_is_present())
                # alert = driver.switch_to.alert
                # builtins.dataframe.loc[soIndx, 'TOOL_REMARKS']='UPLOADED'#alert.text
            else:
                if txtZipped in builtins.dataframe.loc[soIndx, 'TOOL_REMARKS']:
                    builtins.dataframe.loc[soIndx, 'STATUS'] = "EXTERNAL DOC ALREADY AVAILABLE"
                    UpdateToolIFRemarks(BLNum, "EXTERNAL DOC ALREADY AVAILABLE")
                else:
                    builtins.dataframe.loc[soIndx, 'STATUS'] = "MISSING DOCS"
                    UpdateToolIFRemarks(BLNum, "CANT PROCEED")
            # SearchResult=WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="searchResultsListPanel"]'))).text
            # if (SearchResult.find("No Records")>0):
            #    CreateExternalDocs(BLNum,ZippedFileLoc,UploadType,UploadRemarks)

        soIndx += 1

    # post upload process here
    # ManagePostProcessing(BLNum)


def ManageWhsBkg(SONum, eDriver):
    DEST = ""
    DR = ""
    DS = ""
    IsSinglePanel = False
    # whsBkgSO = driver.find_element(By.CLASS_NAME, "slick-cell l3 r3 lni-editable-cell")
    WebDriverWait(eDriver, 6).until(
        EC.element_to_be_clickable((By.XPATH, '//*[@id="criteriaList"]/div[5]/div/div[2]/div[4]'))).click()
    # whsBkgSO.send_keys(SONum)
    ##WebDriverWait(driver, 3).until(
    ##    EC.element_to_be_clickable((By.XPATH, '/html/body/div[4]/textarea'))).send_keys(SONum)
    eDriver.find_element(By.TAG_NAME, 'textarea').send_keys(SONum)
    # click RUN button
    eDriver.execute_script("""top.executeSlickGridFunctions(); parent.gui.screen.run()""")
    time.sleep(6)
    if IsPageReady(eDriver):
        WebDriverWait(eDriver, 4).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '#quickSearchText'))).click()
        try:
            #WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#bookingDestField'))).click()
            Dest = eDriver.find_element(By.CSS_SELECTOR, "#bookingDestField")
            Dest = Dest.get_attribute('value')
            Btn861 = eDriver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/div[1]/div/div[2]/div[2]/div[2]/div[2]/ul/li[2]/a")
            Btn861.click()
            IsSinglePanel = True
            time.sleep(2)
            if IsPageReady(eDriver):
                rctDate = eDriver.find_element(By.CSS_SELECTOR, "#boi1OptionalValHiddenField2")
                DR = rctDate.get_attribute('value')
                # builtins.dataframe['DOC RECEIPT DATE'] = rctDate.get_attribute('value')
                scanDate = eDriver.find_element(By.CSS_SELECTOR, "#boi1OptionalValHiddenField0")
                DS = scanDate.get_attribute('value')
                # builtins.dataframe['DOC SCAN DATE'] = scanDate.get_attribute('value')
                updateBkgDetails(eDriver, SONum, DR, DS, Dest)
        except NoSuchElementException:
            WebDriverWait(eDriver, 3).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, '#quickSearchText'))).click()
            time.sleep(2)

        # if IsPageReady(eDriver):
        if IsSinglePanel==False:
            WebDriverWait(eDriver, 7).until(EC.element_to_be_clickable((By.XPATH,
                                                                 '/html/body/div[1]/div[2]/div[2]/div[1]/div/div[2]/div[2]/div[2]/div[2]/div/div/div/div[2]/div[2]/div/div[1]/span/input'))).click()

        # for ck in chkHdrs:
        #    ck.click()
        # for chkHdrs in driver.find_elements(By.CLASS_NAME, "ui-state-default slick-header-column lni-fixed-header"):
        #    chkHdrs.click()
        # for checkboxes in driver.find_elements(By.CSS_SELECTOR("#resultList > div.slick-viewport > div > div:nth-child(1) > div.slick-cell.l0.r0.slick-cell-checkboxsel > input[type=checkbox]")):
        #    checkboxes.click()
        if IsPageReady(eDriver):
            if IsSinglePanel == False:
                eDriver.execute_script("""top.executeSlickGridFunctions(); parent.gui.screen.openBatch()""")
        #time.sleep(5)
        if IsPageReady(eDriver):
            if IsSinglePanel == False:
                getBkgDetails(eDriver, SONum)
                time.sleep(3)
        # builtins.dataframe['DOC RECEIPT DATE']=GetInputElemValByClass("slick-cell l3 r3 lni-editable-cell")
        # builtins.dataframe['DOC SCAN DATE'] =GetInputElemValByClass("slick-cell l4 r4 lni-editable-cell")
        # builtins.dataframe['DEST'] = GetInputElemValByClass("slick-cell l5 r5")



def GetFirstInputElemByXpathInTbl(fXpath, colHdrName, eDriver):
    EleCols = ""
    FindTxt = ""
    FindTxt = WebDriverWait(eDriver, 20).until(EC.visibility_of_element_located((By.XPATH, fXpath))).text
    return FindTxt


def GetInputElemValByClass(className,eDriver):
    elem = eDriver.find_elements(By.CLASS_NAME, className)
    return elem.text


def GetInputElemValByID(elemID,eDriver):
    elem = eDriver.find_element(By.ID, elemID)
    return elem.text
def GetInputElemValByNAME(elemName,eDriver):
    elem = eDriver.find_element(By.NAME, elemName)
    return elem.text

def IsStrByTextVisibleOnUI(srcStr):
    try:
        srcElement = driver.find_element(By.XPATH(
            '/html/body/div[1]/div[2]/div[2]/div[1]/div/div[2]/div[2]/div[2]/div[2]/div/div/div/div[2]/div[2]/table/tbody/tr[1]/td/div/div'))
        # srcElement = driver.find_element(By.XPATH("//*[contains(text(), " + srcStr + "]"))
        if (srcElement.is_displayed()):
            return True
        else:
            return False
    except Exception as e:
        return False


def IsProceedToExtDocCreation(eDriver):
    chkElement = ""
    proceed = False
    # check presence of Properties Button
    chkElement = 'Properties'
    hasExtDoc = False
    hasExtDoc = IsStrBySelectorVisibleOnUI(chkElement, eDriver)
    if hasExtDoc:
        proceed = False
    else:
        chkElement = 'filter'
        hasExtDoc = IsStrBySelectorVisibleOnUI(chkElement, eDriver)
        if hasExtDoc:
            proceed = False
            return proceed
    if (proceed):
        return proceed
    else:
        chkElement = 'norecords'
        hasExtDoc = IsStrBySelectorVisibleOnUI(chkElement,eDriver)
        if (hasExtDoc):
            proceed = True
        else:
            proceed = False
    # check presence of "No records found"
    return proceed


def IsStrByIDVisibleOnUI(srcStrByID, eDriver):
    elemID = ""
    try:
        if srcStrByID == 'Login':
            elemId = 'errorZone'
            srcElement = eDriver.find_element(By.ID(elemId))
            if (srcElement.is_displayed()):
                return True
            else:
                return False
        if srcStrByID == 'norecords':
            elemId = 'searchResultsListPanel'
            srcElement = eDriver.find_element(By.ID(elemId))
            if (srcElement.is_displayed()):
                return True
            else:
                return False
        if srcStrByID == 'Properties':
            elemId = 'docId'
            srcElement = eDriver.find_element(By.NAME, 'docId')
            if (srcElement.is_displayed()):
                return True
            else:
                return False
    except Exception as e:
        return False


def IsStrBySelectorVisibleOnUI(srcElement, eDriver):
    if srcElement == 'OK':
        try:
            eDriver.find_element(By.CSS_SELECTOR, "body > div.ui-dialog.ui-corner-all.ui-widget.ui-widget-content.ui-front.ui-dialog-buttons.ui-draggable.ui-resizable > div.ui-dialog-buttonpane.ui-widget-content.ui-helper-clearfix > div > button")
        except NoSuchElementException:
            return False
        return True
    if srcElement == 'BkgProperties':
        try:
            eDriver.find_element(By.CSS_SELECTOR, "#shpOptionalValHiddenField1")
        except NoSuchElementException:
            return False
        return True
    if srcElement == 'Login':
        try:
            eDriver.find_element(By.CSS_SELECTOR, "#errorZone")
        except NoSuchElementException:
            return False
        return True
    if srcElement == 'Properties':
        try:
            eDriver.find_element(By.CSS_SELECTOR, "#docDate")
        except NoSuchElementException:
            return False
        return True
    if srcElement == 'norecords':
        try:
            eDriver.find_element(By.CSS_SELECTOR, "#searchResultsListPanel")
        except NoSuchElementException:
            return False
        return True
    if srcElement == 'filter':
        try:
            eDriver.find_element(By.CSS_SELECTOR, "#searchResultsListShowFilterButton > nobr")
        except NoSuchElementException:
            return False
        return True


def IsStrByXPathVisibleOnUI(srcElement, eDriver):
    if srcElement == 'OK':
        try:
            element = eDriver.find_element(By.XPATH('/html/body/div[5]/div[3]/div/button'))
            return True
        except NoSuchElementException:
            return False
    if srcElement == 'filter':
        try:
            element = eDriver.find_element(By.XPATH('//*[@id="searchResultsListShowFilterButton"]/nobr'))
            return True
        except NoSuchElementException:
            return False
    if srcElement == 'BkgProperties':
        try:
            element = eDriver.find_element(
                By.XPATH('//*[@id="sections-0Layer"]/div[1]'))
            return True
        except NoSuchElementException:
            return False
    if srcElement == 'properties':
        try:
            element = eDriver.find_element(
                By.XPATH('/html/body/div[1]/div[2]/div[2]/div[1]/div/div[2]/div[2]/div[2]/div[2]/ul/li[1]'))
            return True
        except NoSuchElementException:
            return False
    if srcElement == 'norecords':
        try:
            element = eDriver.find_element(By.ID('searchResultsListPanel'))
            txtContent = ""
            resContent = ""
            txtContent = element.get_attribute('innerHTML')
            resContent = txtContent.strip()
            if resContent.find('no'):
                return True
            else:
                return False
        except NoSuchElementException:
            return False


def GetUniqueVals(list1):
    list_set = set(list1)
    # convert the set to the list
    unique_vals = (list(list_set))
    return unique_vals

def IsInstr(self, srcKey):
    srchKey = ""
    srchString = ""
    srchKey = srcKey;
    srchString = str(self)
    srcRes = 0
    try:
        if srchKey in self:
            srcRes = self.find(srchKey, 0)
        else:
            srcRes = -1
        return srcRes
    except NoSuchElementException:
        srcRes = -1
    return srcRes


# def ProcessUniqueBLs():
#     for ub in UniqueBL:
#         gCurOnProcBL=ub
#         gLocalForUploadFolder ="C:\\Users\\ignacio.teodorojr\\ADIProjects\\ADIDocUpload\\" + gCurOnProcBL
#         #create temp local for consolidation folder using BL as folder name
#         makeFolder(gLocalForUploadFolder)
#         for i, row in enumerate(builtins.dataframe.values):
#             if (builtins.dataframe.index[i]['BL']) ==gCurOnProcBL:
#                 if (builtins.dataframe.index[i]['STATUS']).find("UPLOADED")==0 and (builtins.dataframe.index[i]['STATUS']).find('COMPLETED')==0:
def IsPWBUploaded(BLNum):
    PWBUploaded = False
    for i, row in enumerate(builtins.dataframe.values):
        if (dataframe.index[i]['BL']) == BLNum:
            if (builtins.dataframe.index[i]['REMARKS']).find("PWB UPLOADED") > 0:
                PWBUploaded = True
    return PWBUploaded


def IsForPartial(BLNum):
    IsForPartial = False
    for i, row in enumerate(builtins.dataframe.values):
        if (builtins.dataframe.index[i]['BL']) == BLNum:
            if (builtins.dataframe.index[i]['STATUS']).find("PARTIAL") > 0:
                IsForPartial = True
    return IsForPartial


def IsJapan(Dest):
    isJapan = False
    if (Dest) == 'JP':
        isJapan = True
    if (Dest) == 'TYO':
        isJapan = True
    if (Dest) == 'TOKYO':
        isJapan = True
    if Dest.find('JP') > 0:
        isJapan = True
    if Dest.find('TYO') > 0:
        isJapan = True
    return isJapan


def ConsolidateFiles(BLNum):
    SrcPath = ""
    DestFolder = ""
    remarks = ""
    soremarks = ""
    zipFolder = ""
    finZipFolder = ""
    uploadtype = ""
    txtNotFound = 'not found'
    txtPartial = 'PARTIAL'
    isJapan = False
    soIndx = 0
    for row in builtins.dataframe.itertuples():
        # BLNum = builtins.dataframe.loc[soIndx, 'BL']
        if row.BL == gCurrBL:
            soIndx = row.Index
            if IsJapan(builtins.dataframe.loc[soIndx, 'DEST']):
                isJapan = True

        DestFolder = ManageDestFolder(BLNum, isJapan)
        uploadtype = builtins.dataframe.loc[soIndx, 'STATUS']
        if (uploadtype.find("PARTIAL")) > 0:
            # perform download of previously upload
            ManagePrevUploadedFiles(BLNum, DestFolder, IsJapan(isJapan))
        soIndx += 1
        # Upload PWB
        DestFolderPath = ""
        if isJapan:
            FileTobeConsolidated = "BL#" + BLNum + ".pdf"
            DestFolderPath = DestFolder + FileTobeConsolidated
        else:
            FileTobeConsolidated = BLNum + ".pdf"
            DestFolderPath = DestFolder + FileTobeConsolidated
        SrcPath = FileLocator(FileTobeConsolidated, BLNum, isJapan)
        if (not (SrcPath)):
            FileTobeConsolidated = BLNum + ".pdf"
            SrcPath = FileLocator(FileTobeConsolidated, BLNum, isJapan)
            if (not (SrcPath)):
                remarks = BLNum + ' not found'
                # if remarks.find("not found"):
            else:
                if isJapan:
                    if not SrcPath:
                        FileTobeConsolidated = "BL#" + BLNum + ".pdf"
                        DestFolderPath = DestFolder + FileTobeConsolidated
                        SrcPath = FileLocator(FileTobeConsolidated, BLNum, isJapan)
        remarks = (Consolidator(BLNum, FileTobeConsolidated, SrcPath, DestFolderPath, isJapan))



        # Upload SO, Invoice, Packing List
        soremarks = UploadSODocs(BLNum, isJapan)

        uploadtype = GetUploadType(BLNum)

        # if txtPartial in uploadtype:
        #     # if txtNotFound in remarks:
        #     #     remarks = 'Consolidated'
        #     # if txtNotFound in soremarks:
        #     #     soremarks = 'Consolidated'
        # else:
        #     if txtNotFound in remarks:
        #         UpdateStatusByBL(BLNum,'PARTIAL UPLOAD')
        # update Location of to be zipped files
        if remarks == 'Consolidated' and soremarks == 'Consolidated':
            zipFolder = DestFolder + '.zip'
            finZipFolder = zipFolder.replace('\\.zip', '.zip')
            zipFolder.replace('\.zip', '.zip')
            UpdateZipFileLocation(BLNum, finZipFolder)
        else:
            if uploadtype=='COMPLETE UPLOAD':
                #UpdateStatusBySO()
                a = "." #to be removed
            else: #do upload for PARTIAL scenario
                if remarks == 'Consolidated' or soremarks == 'Consolidated':
                    zipFolder = DestFolder + '.zip'
                    finZipFolder = zipFolder.replace('\\.zip', '.zip')
                    zipFolder.replace('\.zip', '.zip')
                    UpdateZipFileLocation(BLNum, finZipFolder)
        return remarks + "|" + soremarks

def UploadAllOtherDocs(BLNum, SO, isJapan):
    FileTobeConsolidated = ""
    DestFolderPath = ""
    InvDestFolderPath = ""
    PLDestFolderPath = ""
    fName = ""
    JPprefix = "SO#"
    remarks = ""
    CIKey = "Inv"
    PLKey = "Pack"
    UploadedDocsRemarks = ""
    currSOFile = ""
    soIndx = 0
    srcRes = 0
    i = 0
    soFList = []
    soFList = RecFileLocator(SO)
    if isJapan:
        InvDestFolderPath = builtins.gLocalForUploadFolder + 'BL#' + BLNum + '\\' + 'Commercial Invoice'
        makeFolder(InvDestFolderPath)
        PLDestFolderPath = builtins.gLocalForUploadFolder + 'BL#' + BLNum + '\\' + 'Packing List'
        makeFolder(PLDestFolderPath)
        for i in range(len(soFList)):
            currSOFile = soFList[i]
            fName = Path(currSOFile).name
            srcRes = IsInstr(str(fName), str(JPprefix))
            if srcRes < 0:
                InvDestFolderPath = InvDestFolderPath + "\\" + "SO#" + fName
            srcRes = IsInstr(str(currSOFile), str(CIKey))
            if not IsDirExists(currSOFile):
                if srcRes > 0:
                    moveFile(currSOFile, InvDestFolderPath)
                    UploadedDocsRemarks = "CI Done"
                srcRes = 0
                srcRes = IsInstr(PLKey, currSOFile)
                if srcRes > 0:
                    srcRes = IsInstr(JPprefix, fName)
                    if srcRes < 0:
                        PLDestFolderPath = PLDestFolderPath + "\\" + "SO#" + fName
                    moveFile(soFList[i], PLDestFolderPath)
                    if not UploadedDocsRemarks:
                        UploadedDocsRemarks = "PL Done"
                    else:
                        UploadedDocsRemarks = UploadedDocsRemarks + "PL Done"
            else:
                #for Inv
                srcRes = IsInstr(str(currSOFile), str(CIKey))
                if srcRes > 0:
                    moveFile(soFList[i], InvDestFolderPath)
                    if not UploadedDocsRemarks:
                        UploadedDocsRemarks = "CI Done"
                    else:
                        UploadedDocsRemarks = UploadedDocsRemarks + "CI Done"
                else:
                    #for PL'
                    srcRes = IsInstr(str(currSOFile), str(PLKey))
                    if srcRes > 0:
                        PLDestFolderPath = PLDestFolderPath + "\\" + "SO#" + fName
                        moveFile(soFList[i], PLDestFolderPath)
                        if not UploadedDocsRemarks:
                            UploadedDocsRemarks = "PL Done"
                        else:
                            UploadedDocsRemarks = UploadedDocsRemarks + ",PL Done"
                    else:
                        #for all other Docs
                        if isJapan:
                            DestFolderPath = builtins.gLocalForUploadFolder + 'BL#' + BLNum + '\\' + fName
                        else:
                            DestFolderPath = builtins.gLocalForUploadFolder + BLNum + '\\' + fName
                        moveFile(soFList[i], DestFolderPath)
                        if not UploadedDocsRemarks:
                            UploadedDocsRemarks = "OTH Done"
                        else:
                            UploadedDocsRemarks = UploadedDocsRemarks + "OTH Done"
        return UploadedDocsRemarks
    else:
        for i in range(len(soFList)):
            currSOFile = soFList[i]
            fName = Path(currSOFile).name
            DestFolderPath = builtins.gLocalForUploadFolder + BLNum + "\\" + fName
            # makeFolder(DestFolderPath)
            if not IsDirExists(DestFolderPath):
                moveFile(currSOFile, DestFolderPath)
                UploadedDocsRemarks = "SO Done"
            else:
                UploadedDocsRemarks = "SO Done"

        return UploadedDocsRemarks
def UploadSODocs(BLNum, isJapan):
    FileTobeConsolidated = ""
    remarks = ""
    soUploadRemarks = ""
    CIRem = "CI Done"
    PLRem = "PL Done"
    SORem = "SO Done"
    soIndx = 0
    soFList = []
    for row in builtins.dataframe.itertuples():
        if row.BL == gCurrBL:
            soIndx = row.Index
            SO = builtins.dataframe.loc[soIndx, 'SO']
            soUploadRemarks = UploadAllOtherDocs(BLNum, SO, isJapan)
            if CIRem in soUploadRemarks and PLRem in soUploadRemarks or SORem in soUploadRemarks:
                soUploadRemarks = "Consolidated"
                remarks = "Consolidated"
            else:
                if isJapan:
                    if not CIRem in soUploadRemarks:
                        # locate Invoice doc
                        DestFolderPath = builtins.gLocalForUploadFolder + 'BL#' + BLNum + '\\' + 'Commercial Invoice'
                        makeFolder(DestFolderPath)
                        FileTobeConsolidated = 'SO#' + str(SO) + '-Invoice.pdf'
                        SrcPath = FileLocator(FileTobeConsolidated, BLNum, isJapan)
                        if SrcPath:
                            remarks = (Consolidator(BLNum, FileTobeConsolidated, SrcPath,
                                                    DestFolderPath + "\\" + FileTobeConsolidated, isJapan))
                        else:
                            FileTobeConsolidated = str(SO) + '-Invoice.pdf'
                            if IsDirExists(DestFolderPath + "\\" + FileTobeConsolidated):
                                remarks = "Consolidated"
                            else:
                                SrcPath = FileLocator(FileTobeConsolidated, BLNum, isJapan)
                                if SrcPath:
                                    remarks = (Consolidator(BLNum, FileTobeConsolidated, SrcPath,
                                                                DestFolderPath + "\\" + FileTobeConsolidated, isJapan))
                                    remarks = "Consolidated"
                                else:
                                    if isJapan:
                                        FileTobeConsolidated = "BL#" + str(BLNum) + "\\" + "\\" + str(SO) + "-Invoice.pdf"
                                        SrcPath = FileLocator(FileTobeConsolidated, BLNum, isJapan)
                                        if SrcPath:
                                            remarks = (Consolidator(BLNum, FileTobeConsolidated, SrcPath,
                                                                    DestFolderPath + "\\" + FileTobeConsolidated, isJapan))
                                        else:
                                            FileTobeConsolidated = "BL#" + str(BLNum) + "\\" + "Commercial Invoice" + "\\" + str(SO) + "-Invoice.pdf"
                                            SrcPath = FileLocator(FileTobeConsolidated, BLNum, isJapan)
                                            if SrcPath:
                                                if IsDirExists(FileTobeConsolidated):
                                                    remarks = "Consolidated"
                                                else:
                                                    remarks = (Consolidator(BLNum, FileTobeConsolidated, SrcPath,
                                                                            DestFolderPath + "\\" + FileTobeConsolidated, isJapan))
                                            else:
                                                remarks = str(BLNum) + " docs not found"
                    # packing list doc
                    if not PLRem in soUploadRemarks:
                        DestFolderPath = builtins.gLocalForUploadFolder + 'BL#' + str(BLNum) + '\\' + 'Packing List'
                        makeFolder(DestFolderPath)
                        FileTobeConsolidated = 'SO#' + str(SO) + '-Packing List.pdf'
                        if IsDirExists(DestFolderPath + "\\" + FileTobeConsolidated):
                            remarks = "Consolidated"
                        else:
                            SrcPath = FileLocator(FileTobeConsolidated, BLNum, isJapan)
                            if SrcPath:
                                remarks = (Consolidator(BLNum, FileTobeConsolidated, SrcPath,
                                                    DestFolderPath + "\\" + FileTobeConsolidated, isJapan))
                            else:
                                FileTobeConsolidated = str(SO) + '-Packing List.pdf'
                                SrcPath = FileLocator(FileTobeConsolidated, BLNum, isJapan)
                                if SrcPath:
                                    remarks = (Consolidator(BLNum, FileTobeConsolidated, SrcPath,
                                                        DestFolderPath + "\\" + FileTobeConsolidated, isJapan))
                                else:
                                    FileTobeConsolidated = "BL#" + str(BLNum) + "\\" + "Packing List" + "\\" + str(SO) + "-Packing List.pdf"
                                    SrcPath = FileLocator(FileTobeConsolidated, BLNum, isJapan)
                                if SrcPath:
                                    if(IsDirExists(FileTobeConsolidated)):
                                        remarks = "Consolidated"
                                    else:
                                        remarks = (Consolidator(BLNum, FileTobeConsolidated, SrcPath,
                                                            DestFolderPath + "\\" + FileTobeConsolidated, isJapan))
                else:
                    DestFolderPath = builtins.gLocalForUploadFolder + str(BLNum)
                    makeFolder(DestFolderPath)
                    FileTobeConsolidated = str(SO) + '.pdf'
                    SrcPath = FileLocator(FileTobeConsolidated, BLNum, isJapan)
                    DestFolderPath = str(DestFolderPath) + '\\' + str(FileTobeConsolidated)
                    if SrcPath:
                        remarks = (Consolidator(BLNum, FileTobeConsolidated, SrcPath, DestFolderPath, isJapan))
                    else:
                        remarks = "SO not found"
                        UpdateStatusBySO(SO, 'PARTIAL UPLOAD')
        soIndx += 1
    return remarks


def ZipTheDocs(BLNum):
    zipFileLoc = ""
    soIndx = 0
    for row in builtins.dataframe.itertuples():  # ['BL']:
        if row.BL == gCurrBL:
            soIndx = row.Index
            if builtins.dataframe.loc[soIndx, 'TOOL_REMARKS'] == 'Consolidated|Consolidated':
                zipFileLoc = builtins.dataframe.loc[soIndx, 'UPLOADED_FILE']
                targetDir = zipFileLoc.replace('.zip', '')
                zipfoldername = os.path.basename(zipFileLoc)  # zipFileLoc.replace('.zip', '')
                rawToZipFile = zipfoldername.replace('.zip', '')
                zippedFileName = str(gCurrBL) + 'zip'
                if FolderZipper(targetDir, zipFileLoc, rawToZipFile):
                    UpdateToolIFRemarks(row.BL, 'Zipped')
                # sys.exit()
                # FileZipper(rawToZipFile,zipFileLoc)
                # zipToUploadFolder(gCurrBL,'zip',rawToZipFile,  zipFileLoc)
                # zipIt(zipFileLoc,zippedFileLoc)

    soIndx += 1
def ManageDestFolder(BLNum, isJapan):
    destFPath = ""
    if not (isJapan):
        destFPath = builtins.gLocalForUploadFolder + BLNum + "\\"
    else:
        destFPath = builtins.gLocalForUploadFolder + "BL#" + BLNum + "\\"
    if not (IsDirExists(destFPath)):
        makeFolder(destFPath)
    return destFPath


def ManagePrevUploadedFiles(BLNum, DestPath, isJapan):
    prevZipPath = ""
    zipFname = BLNum + ".zip"
    zipFpath = builtins.gCompleteFolder + "\\" + zipFname
    prevZipPath = FileLocator(zipFpath, BLNum, isJapan)
    if (prevZipPath):
        unzipIt(prevZipPath, DestPath)


def Consolidator(BLNum, FileTobeConsolidated, SrcFolderPath, DestFolderPath, isJapan):
    if (SrcFolderPath):
        if IsInstr(SrcFolderPath, '\\\\') != -1:
            SrcFolderPath = SrcFolderPath.replace('\\\\', '\\')
    if IsInstr(DestFolderPath, '\\\\') != -1:
        DestFolderPath = DestFolderPath.replace('\\\\', '\\')
    remarks = ""
    fileFound = False
    isFileAvalable = False
    CurrFileToBeIncluded = ""
    SrcPath = ""
    if SrcFolderPath == DestFolderPath:
        remarks = "Consolidated"
        return remarks
    # find FileTobeConsolidated in local to be uploaded folder
    if not (SrcFolderPath):
        SrcFolderPath = FileLocator(FileTobeConsolidated, BLNum, isJapan)
    if (SrcFolderPath):
        if(IsDirExists(DestFolderPath)):
            remarks = "Consolidated"
        else:

            moveFile(SrcFolderPath, DestFolderPath)
            remarks = "Consolidated"
    else:
        remarks = FileTobeConsolidated + " not found"

    return remarks


def RecFileLocator(Fname):
    list = []
    name_path = ""
    filename = ""
    substring = ""
    wchar = "*"
    # #recursively find files in FC Folder which has the same SO# on its fileneme
    SrcDir = builtins.gFCFolder
    substring = str(wchar) + str(Fname) + str(wchar)
    for root, subdirs, files in os.walk(SrcDir):
        for filename in files:
            if str(Fname) in filename:
                name_path = os.path.join(root, filename)
                list.insert(len(list), name_path)
    SrcDir = builtins.gLocalForUploadFolder
    for path in Path(SrcDir).glob("*" + str(Fname) + "*"):
        print(path)
        if str(Fname) in filename:
            name_path = os.path.join(root, filename)
            list.insert(len(list), name_path)
    #recursively find files in locla for upload folder which has the same SO# on its fileneme
    for root, subdirs, files in os.walk(SrcDir):
        for filename in files:
            if str(Fname) in filename:
                name_path = os.path.join(root, filename)
                list.insert(len(list), name_path)
    return list


def FileLocator(FileTobeLocated, BLNum, isJapan):
    fileFound = False
    SrcPath = ""
    # find FileTobeConsolidated in local to be uploaded folder
    SrcPath = builtins.gLocalForUploadFolder + FileTobeLocated  # check if its already available on local to upload folder
    SrcPath.replace("\\\\", "\\")
    isFileAvalable = IsDirExists(SrcPath)
    if isFileAvalable:
        return SrcPath
    else:
        SrcPath = builtins.gLocalForUploadFolder + "\\" + BLNum + "\\" + FileTobeLocated  # check if its already available on local to upload folder
        isFileAvalable = IsDirExists(SrcPath)
        if isFileAvalable:
            return SrcPath
        else:
            SrcPath = builtins.gFCFolder + "\\" + FileTobeLocated
            isFileAvalable = IsDirExists(SrcPath)
            if isFileAvalable:
                return SrcPath
            else:
                SrcPath = builtins.gFCFolder + "\\" + BLNum + "\\" + FileTobeLocated  # check if it's available on FC folder under teamsite
                isFileAvalable = IsDirExists(SrcPath)
                if isFileAvalable:
                    return SrcPath
                else:
                    SrcPath = builtins.gLocalForUploadFolder + "\\" + "BL#" + BLNum + "\\" + FileTobeLocated  # check if its already available on local to upload folder
                    isFileAvalable = IsDirExists(SrcPath)
                    if isFileAvalable:
                        return SrcPath


def JPConsolidator(DestFolderPath, RefNum, uStatus):
    fileFound = False
    isPartial = False
    isFileAvalable = False
    CurrFileToBeIncluded = ""
    CurrFileToBeIncluded = RefNum + ".pdf"
    SrcPath = ""
    SrcPath = builtins.gLocalForUploadFolder + "\\" + RefNum + "\\" + CurrFileToBeIncluded  # check if its already available on local to upload folder
    if not (IsPWBUploaded(RefNum)):
        isPartial = IsForPartial(RefNum)
        isFileAvalable = IsDirExists(SrcPath)
        if not (isFileAvalable):
            SrcPath = builtins.gFCFolder + "\\" + CurrFileToBeIncluded
            isFileAvalable = IsDirExists(builtins.gFCFolder + "\\" + CurrFileToBeIncluded)
        else:
            moveFile(SrcPath, DestFolderPath)
            fileFound = True
def UpdateSinglSOtoDF(SO):
    soIndx = 0
    i = 0
    for row in builtins.dataframe.itertuples():
        if row.BL == gCurrBL:
            soIndx = row.Index
            dfLen = 0
            dfLen = len(builtins.dataframe)
            if soIndx > dfLen:
                builtins.dataframe.loc[builtins.dataframe.shape[0], ['BL']] = builtins.dataframe.index[soIndx]['BL']
                builtins.dataframe.loc[builtins.dataframe.shape[0], ['DOC RECEIPT DATE']] = builtins.dataframe.index[soIndx][
                    'DOC RECEIPT DATE']
                builtins.dataframe.loc[builtins.dataframe.shape[0], ['DOC SCAN DATE']] = builtins.dataframe.index[soIndx][
                    'DOC SCAN DATE']
                builtins.dataframe.loc[builtins.dataframe.shape[0], ['SO']] = SO
                builtins.dataframe.loc[builtins.dataframe.shape[0], ['DEST']] = builtins.dataframe.index[soIndx]['DEST']
            else:
                if not (row.SO):
                    # builtins.dataframe.loc[builtins.dataframe.shape[0], ['BL']] = builtins.dataframe.index[i]['BL']
                    # builtins.dataframe.loc[builtins.dataframe.shape[0], ['DOC RECEIPT DATE']] = builtins.dataframe.index[i]['DOC RECEIPT DATE']
                    # builtins.dataframe.loc[builtins.dataframe.shape[0], ['DOC SCAN DATE']] = builtins.dataframe.index[i]['DOC SCAN DATE']
                    # builtins.dataframe.loc[builtins.dataframe.shape[0], ['SO']] = SO
                    # builtins.dataframe.loc[builtins.dataframe.shape[0], ['DEST']] = builtins.dataframe.index[i]['DEST']
                # else:
                    builtins.dataframe.loc[soIndx, 'SO'] = SO


def UpdateSOtoDF():
    dfCurrBL = ""
    dfCurrSO = ""
    soIndx = 0
    currSOindxPref = 0
    dfLastIndx = 0
    newso = ""
    l = 0
    dfLastIndx = len(builtins.dataframe)
    for i, so in enumerate(builtins.VendorBkg.itertuples(), 1):
        newso = so[1]
        found = False
        for row in builtins.dataframe.itertuples():
            if row.BL == gCurrBL:
                if row.SO == newso:
                    found = True

        if not(found):
            for row in builtins.dataframe.itertuples():
                if row.BL == gCurrBL:
                    if row.SO == '':
                        soIndx = row.Index
                        builtins.dataframe.loc[soIndx, 'SO'] = newso
                        found = True
            if not (found):
                currSOindxPref = AddSORowtoDF(newso, currSOindxPref)

        soIndx += 1

def AddSORowtoDF(SONum, soIndxPref):
    dfCurrBL = ""
    dfCurrSO = ""
    soIndx = 0
    dfLastIndx = 0
    found = False
    l = 0
    dfLastIndx = len(builtins.dataframe)
    # initialize index prefix
    if soIndxPref == 0:
        soIndxPref = 0.1
    else:
        soIndxPref = soIndxPref + 0.1
    dfColCount = 0
    dfColCount = len(builtins.dataframe.axes[1])
    for row in builtins.dataframe.itertuples():
        if row.BL == gCurrBL:
            if row.SO == SONum:
                found = True

        if not (found):
            for row in builtins.dataframe.itertuples():
                if row.BL == gCurrBL:
                    if row.SO == '':
                        soIndx = row.Index
                        builtins.dataframe.loc[soIndx, 'SO'] = SONum
                        found = True
                        return
            if not (found):
                for row in builtins.dataframe.itertuples():
                    if row.BL == gCurrBL:
                        soIndx = row.Index + soIndxPref
                        if dfColCount==9:
                            builtins.dataframe.loc[soIndx] = [gCurrBL, SONum, '.', '.', '.', '.', '.', '.', '.']
                        if dfColCount==11:
                            builtins.dataframe.loc[soIndx] = [gCurrBL, SONum, '.', '.', '.', '.', '.', '.', '.', '.', '.']
                        builtins.dataframe.sort_index(axis=0, ascending=True) #, inplace=False, kind='quicksort')
                        found = True
                        return soIndxPref
                    # builtins.dataframe.loc[builtins.dataframe.shape[0], ['SO']] = newso
                    # dfLastIndx = len(builtins.dataframe) - 1
                    # builtins.dataframe.loc[dfLastIndx, 'BL'] = gCurrBL
                    # builtins.dataframe.loc[dfLastIndx, 'DOC RECEIPT DATE'] = '.'
                    # builtins.dataframe.loc[dfLastIndx, 'DOC SCAN DATE'] = '.'
                    # builtins.dataframe.loc[dfLastIndx, 'DEST'] = '.'

        soIndx += 1


def UpdateZipFileLocation(BLNum, ZFLocation):
    soIndx = 0
    ZFLocation.replace('\\.zip', '.zip')
    for row in builtins.dataframe.itertuples():  # ['BL']:
        if row.BL == gCurrBL:
            soIndx = row.Index
            builtins.dataframe.loc[soIndx, 'UPLOADED_FILE'] = ZFLocation
        soIndx += 1

def UpdateStatusBySO(SONum,Status):
    soIndx = 0
    for row in builtins.dataframe.itertuples():  # ['SO']:
        soIndx = row.Index
        if builtins.dataframe.loc[soIndx, 'SO'] == SONum:
            builtins.dataframe.loc[soIndx, 'STATUS'] = Status
            # if Remarks == 'Consolidated|Consolidated':
            #    builtins.dataframe.loc[soIndx, 'STATUS'] = 'COMPLETE UPLOAD'
        soIndx += 1
def UpdateStatusByBL(BLNum,Status):
    soIndx = 0
    for row in builtins.dataframe.itertuples():  # ['BL']:
        if row.BL == BLNum:
            soIndx = row.Index
            builtins.dataframe.loc[soIndx, 'STATUS'] = Status
            # if Remarks == 'Consolidated|Consolidated':
            #    builtins.dataframe.loc[soIndx, 'STATUS'] = 'COMPLETE UPLOAD'
        soIndx += 1
def GetUploadType(BLNum):
    UploadTyp = ""
    soIndx = 0
    for row in builtins.dataframe.itertuples():
        # BLNum = builtins.dataframe.loc[soIndx, 'BL']
        if row.BL == gCurrBL:
            soIndx = row.Index
            UploadTyp = builtins.dataframe.loc[soIndx, 'STATUS']
    return UploadTyp
def UpdateToolRemarks(BLNum, Remarks):
    soIndx = 0
    for row in builtins.dataframe.itertuples():  # ['BL']:
        if row.BL == BLNum:
            soIndx = row.Index
            builtins.dataframe.loc[soIndx, 'TOOL_REMARKS'] = Remarks
            #if Remarks == 'Consolidated|Consolidated':
            #    builtins.dataframe.loc[soIndx, 'STATUS'] = 'COMPLETE UPLOAD'
        soIndx += 1
# This will update Remarks to TOOL REMARKS column in Main Sht
def UpdateToolIFRemarks(BLNum, Remarks):
    soIndx = 0
    for row in builtins.dataframe.itertuples():
        if row.BL == BLNum:
            soIndx = row.Index
            builtins.dataframe.loc[soIndx, 'TOOL_REMARKS'] = builtins.dataframe.loc[soIndx, 'TOOL_REMARKS'] + '|' + Remarks
        soIndx += 1
def CheckRemarksIfComplete(BLNum):
    found = False
    strUploaded = "Uploaded"
    strCompleted = "Available"
    soIndx = 0
    for row in builtins.dataframe.itertuples():
        if row.BL == BLNum:
            soIndx = row.Index
            if strUploaded in(builtins.dataframe.loc[soIndx, 'TOOL_REMARKS'] ):
                found=True
            if strCompleted in(builtins.dataframe.loc[soIndx, 'TOOL_REMARKS'] ):
                found=True
        soIndx += 1
    return found
def DFcleanUp():
    # cleanup df headers to take spaces
    builtins.dataframe.columns = [c.replace(' ', '_') for c in builtins.dataframe.columns]
    for row in builtins.dataframe.itertuples():  # ['BL']:
        if row.BL == gCurrBL:
            soInBL += 1


def CheckUploadType():
    soInBL = 0
    ctOfNoDR = 0
    ctOfNoDS = 0
    uploadType = ""
    ##count number of SO included in the current BL
    # for i, row in enumerate(builtins.dataframe.values):
    # cleanup df headers to take spaces
    DFcleanUp
    ## builtins.dataframe.columns = [c.replace(' ', '_') for c in builtins.dataframe.columns]
    for row in builtins.dataframe.itertuples():  # ['BL']:
        if row.BL == gCurrBL:
            soInBL += 1

    # for 1SO 1 BL
    if soInBL == 1:
        uploadType = "COMPLETE UPLOAD"
        # soIndx = 0
        # numofEmpty = 0
        # cDR = ""
        # cDS = ""
        # i = 0
        # for i, row in enumerate(builtins.dataframe.values):
        #     if (builtins.dataframe.loc[i, 'BL']) == gCurrBL:
        #         # cDR = builtins.dataframe.loc[i, 'DOC_RECEIPT_DATE']
        #         # cDS = builtins.dataframe.loc[i, 'DOC_SCAN_DATE']
        #         cDR = builtins.dataframe["DOC_RECEIPT_DATE"].values[0]
        #         ## cDR = builtins.dataframe['DOC_RECEIPT_DATE'].iloc[i]
        #         cDS = builtins.dataframe['DOC_SCAN_DATE'].iloc[i]
        #         cDDR = ""
        #         cDDR = cDR[0]
        #         #cDR = builtins.dataframe.at[builtins.dataframe.index[i], 'DOC_RECEIPT_DATE']
        #         #cDS = builtins.dataframe.at[builtins.dataframe.index[i], 'DOC_SCAN_DATE']
        #         ##cDS = builtins.dataframe["DOC_SCAN_DATE"].values[0]
        #         if cDR == ".":
        #             numofEmpty += 1
        #         if cDS == ".":
        #             numofEmpty += 1
        #         if builtins.dataframe.loc[i, 'DOC_SCAN_DATE'] == ".":
        #             numofEmpty += 1
        #         if numofEmpty > 0:
        #             uploadType = "PARTIAL UPLOAD"
        #         else:
        #             uploadType = "COMPLETE UPLOAD"
        #
        #         # if (builtins.dataframe.loc[soIndx, 'DOC_RECEIPT_DATE'] and builtins.dataframe.loc[soIndx, 'DOC_SCAN_DATE']):
        #         #     uploadType = "COMPLETE UPLOAD"
        #         # else:
        #         #     uploadType = "PARTIAL UPLOAD"
    else:
        soIndx = 0
        for row in builtins.dataframe.itertuples():
            if row.BL == gCurrBL:
                soIndx = row.Index
                if not (builtins.dataframe.loc[soIndx, 'DOC_RECEIPT_DATE']):
                    ctOfNoDR += 1
                if not (builtins.dataframe.loc[soIndx, 'DOC_SCAN_DATE']):
                    ctOfNoDS += 1
            soIndx += 1

        if ctOfNoDR > 0:
            uploadType = "PARTIAL UPLOAD"
        if ctOfNoDR == 0:
            if ctOfNoDS == 0:
                uploadType = "COMPLETE UPLOAD"
        else:
            if ctOfNoDS < soInBL:
                uploadType = "PARTIAL FOR COMPLETE UPLOAD"

    # update Upload Type in builtins.dataframe
    soIndx = 0
    for row in builtins.dataframe.itertuples():
        if row.BL == gCurrBL:
            soIndx = row.Index
            builtins.dataframe.loc[soIndx, 'STATUS'] = uploadType
            soIndx += 1
    return uploadType

def main():
    globalInitializer()
    if len(sys.argv) > 1:
        in_username = sys.argv[1]
        print('username', in_username)
        in_userpwd = sys.argv[2]
        print('userpwd', in_userpwd)
        in_xlToolPath = sys.argv[3]
        print(in_xlToolPath)
    else:
        os.system('cls')
        in_username = input("Pls enter your MyPA Username: ")
        in_userpwd = pwinput.pwinput(prompt='Pls enter your MyPA Password:  ', mask='*')

    # attach to existing xl---------------
    # wb1=win32com7.client.GetObject(str(gMainWbk))
    # wb1.Application.Run("DocUploadTool 003.003.xlsb!WorkList.ClearWorkList")
    # xl=win32com.client.Dispatch("Excel.Application")
    # ---------------------------
    # excel = win32.gencache.EnsureDispatch('Excel.Application')
    # excel.Application.Run("DocUploadTool 003.003.xlsb!WorkList.ClearWorkList")
    # for wb in excel.Workbooks:
    #     print(wb.Name)
    #     excel.Application.run
    #     if(wb.Name==gMainWbk):
    #         print (wb.Name)

    # builtins.dataframe = pd.read_excel(gMainWbk, skiprows=5, usecols='A:N', engine='pyxlsb', sheet_name='Main')  # , column="I", row=2)
    # xls = pd.ExcelFile('path_to_file.xls')
    builtins.dataframe = pd.read_excel(builtins.g2bProcXl, 'Main')
    builtins.dataframe['SO'] = ""
    # builtins.dataframe = pd.read_excel(builtins.g2bProcXl, usecols='A:F', sheet_name='Main')  # , column="I", row=2)

    # builtins.dataframe =pd.read_excel(r"C:\Users\ignacio.teodorojr\ADIProjects\ADIDocUpload\DocUploadTool_001.3.4.6.61.83.2.xlsb", skiprows=5, usecols='A:N', engine='pyxlsb',sheet_name='Main')#, column="I", row=2)
    # print(builtins.dataframe)

    RowCount = 0

    # workArr=np.array(['BL','SO','DOC RECEIPT DATE','DOC UPLOAD DEADLINE','STATUS','REMARKS','DATE UPLOADED','PIC','FOR UPLOAD SRC','DOC SCAN DATE','DEST','TOOL REMARKS','COMPLETE?','UPLOADED FILE'])
    headersList = ["BL", "SO", "DOC RECEIPT DATE", "DOC UPLOAD DEADLINE", "STATUS", "REMARKS", "DATE UPLOADED", "PIC",
                   "FOR UPLOAD SRC", "DOC SCAN DATE", "DEST", "TOOL REMARKS", "COMPLETE?", "UPLOADED FILE"]

    global WorkList
    WorkList = []
    for column in builtins.dataframe['BL']:
        # columnSeriesObj = builtins.dataframe[column]
        cellvalue = column
        print(cellvalue)
        if not (pd.isnull(cellvalue)):

            if not (cellvalue in WorkList):
                WorkList.append(cellvalue)
                RowCount += 1

        print('Column Name : ', column)
        print('Column Contents : ', cellvalue)

    if not in_username:
        in_username = input("Username ")  # 'IGTEODORO'

    if not in_userpwd:
        in_userpwd = input("Password ")

    edge_options = EdgeOptions()
    edge_options.use_chromium = True
    # edge_options.add_argument("C:\\Users\\"+ gCurruser +"\\AppData\\Local\\Microsoft\\Edge\\User Data")
    edge_options.add_argument("--user-data-dir=C:\\Users\\" + builtins.gCurruser + "\\AppData\\Local\\Microsoft\\Edge\\Work")
    edge_options.add_argument("--start-maximized")
    edge_options.add_argument("_ignore_local_proxy")
    # driver = Edge(builtins.gEdgeDriver, options=edge_options)
    # driver = webdriver.Edge(builtins.gEdgeDriver, options=edge_options)
    # driver = webdriver.Edge('C:\\Users\\" + gCurruser + "\\ADIDocUpload\\msedgedriver.exe')

    driver = webdriver.Edge('C:\\Users\\' + builtins.gCurruser + '\\ADIDocUpload\\msedgedriver.exe')
    _url = driver.command_executor._url
    _sessionID = driver.session_id
    # driver=webdriver.remote(command_executor=_url,desired_capabilities={})
    # driver.close()
    # driver.session_id=_sessionID
    # trUrl="panalpina.log-net.com"
    # print(_url.find(trUrl))
    # if not _url:
    #     driver = webdriver.Edge('C:\\Users\\ignacio.teodorojr\\ADIProjects\\ADIDocUpload\\venv\\msedgedriver.exe')

    # driver.get('https://panalpina.log-net.com/')
    driver.get(builtins.gMyPAURL)

    time.sleep(3)
    mBWindowHndl = driver.current_window_handle
    try:
        # WebDriverWait(driver, 3).until(EC.text_to_be_present_in_element_value(By.ID,"checkLoginForm"))
        username = driver.find_element(By.NAME, "userid")
        username.send_keys(in_username)

        xpassword = driver.find_element(By.NAME, "password")
        xpassword.send_keys(in_userpwd)

        # loginbutton = driver.find_element(By.CLASS_NAME, "m-quick-access-widget__button js-m-quick-access-widget__button")
        driver.execute_script('return checkLoginForm()')
        time.sleep(7)
        # catch an Internal Lognet error prompt. This notification is displayed when there's a downtime in Lognet
        try:
            # driver.find_element(By.CSS_SELECTOR, "body > div.ui-dialog.ui-corner-all.ui-widget.ui-widget-content.ui-front.ui-draggable.ui-resizable > div.ui-dialog-titlebar.ui-corner-all.ui-widget-header.ui-helper-clearfix.ui-draggable-handle > button > span.ui-button-icon.ui-icon.ui-icon-closethick")
            # driver.find_element(By.CSS_SELECTOR,
            # "body > div.ui-dialog.ui-corner-all.ui-widget.ui-widget-content.ui-front.ui-draggable.ui-resizable > div.ui-dialog-titlebar.ui-corner-all.ui-widget-header.ui-helper-clearfix.ui-draggable-handle > button > span.ui-button-icon.ui-icon.ui-icon-closethick").click()
            driver.find_element(By.XPATH, "//button[@title='Close']").click()

            handles = driver.window_handles
            javaErrTitle = "Java Error"
            for i in handles:

                # close specified web page
                if javaErrTitle in driver.title:
                    driver.switch_to.window(i)
                    time.sleep(2)
                    driver.close()
        except NoSuchElementException:
            time.sleep(1)
        try:
            driver.find_element(By.CSS_SELECTOR, "#errorZone")
            Mbox('ADI Doc Upload Tool v.7.2', 'Login Failed! Pls enter correct MyPA credentials', 0)
            return
        except NoSuchElementException:
            # catch an Internal Lognet error prompt. This notification is displayed when there's a downtime in Lognet
            driver.switch_to.frame('viewFrame')
            try:
                close_buttons = WebDriverWait(driver, 2).until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR,
                                                         'body > div.ui-dialog.ui-corner-all.ui-widget.ui-widget-content.ui-front.ui-draggable.ui-resizable > div.ui-dialog-titlebar.ui-corner-all.ui-widget-header.ui-helper-clearfix.ui-draggable-handle > button')))
                # check how many buttons in on the HTML, you can try "visibility_of_all_elements_located"
                print(len(close_buttons))

                visible_buttons = [close_button for close_button in close_buttons if close_button.is_displayed()]
                visible_buttons_len = len(visible_buttons)

                visible_buttons[visible_buttons_len - 1].click()
                searchType = driver.find_element(By.ID, "quickSearchType-button")
                searchType.click()
                searchRef = driver.find_element(By.ID, "ui-id-14")
                searchRef.click()  # .send_keys("Shipments by B/L")
            except NoSuchElementException:

                WebDriverWait(driver, 120).until(
                    EC.frame_to_be_available_and_switch_to_it((By.XPATH, '/html/frameset/frame[1]')))
                # driver.switch_to.frame('viewFrame')
                searchType = driver.find_element(By.ID, "quickSearchType-button")
                searchType.click()
                searchRef = driver.find_element(By.ID, "ui-id-13")
                searchRef.click()  # .send_keys("Shipments by B/L")
    except TimeoutException:
        print("Loading took too much time!")
    # def WaitUntilLoadComplete():

    global gCurrBL
    gCurrBL = ""
    remarks = ""
    searchType = driver.find_element(By.ID, "quickSearchType-button")
    searchType.click()
    searchRef = driver.find_element(By.ID, "ui-id-13")
    searchRef.click()  # .send_keys("Shipments by B/L")
    DFcleanUp
    # loop through list
    builtins.dataframe['DOC_RECEIPT_DATE'] = "."
    builtins.dataframe['DOC_SCAN_DATE'] = "."
    builtins.dataframe['DEST'] = "."

    wlCt = 0
    for wlCt in range(len(WorkList)):
        searchText = driver.find_element(By.ID, "quickSearchText")
        WebDriverWait(driver, 4).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '#quickSearchText'))).clear()
        searchText.send_keys(WorkList[wlCt])
        searchButton = driver.find_element(By.ID, "quickSearchButton")
        searchButton.click()
        time.sleep(6)
        gCurrBL = WorkList[wlCt]

        try:
            WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#shpOptionalValHiddenField1'))).click()
            VBkg = driver.find_element(By.CSS_SELECTOR, "#shpOptionalValHiddenField1")
            builtins.VendorBkg = VBkg.get_attribute('value')
            UpdateSinglSOtoDF(builtins.VendorBkg)
            builtins.SOCount = 1
            builtins.gSOGrpType = "Single"
        except NoSuchElementException:
            WebDriverWait(driver, 3).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, '#quickSearchText'))).click()
            # GetMultipleSO(driver)
            # builtins.VendorBkg = getVdrBkgMultiple(driver)
            builtins.VendorBkg = GetMultipleSO(driver)
            UpdateSOtoDF()
            builtins.gSOGrpType = "Multiple"
            # WebDriverWait(driver, 120).until(
            #     EC.frame_to_be_available_and_switch_to_it((By.XPATH, '/html/frameset/frame[1]')))
            # # driver.switch_to.frame('viewFrame')
            # searchType = driver.find_element(By.ID, "quickSearchType-button")
            # searchType.click()
            # searchRef = driver.find_element(By.ID, "ui-id-13")
            # searchRef.click()  # .send_keys("Shipments by B/L")
        except TimeoutException:
            WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#quickSearchText'))).click()
            builtins.VendorBkg = GetMultipleSO(driver)
            UpdateSOtoDF()
            builtins.gSOGrpType = "Multiple"
            # builtins.VendorBkg = getVdrBkgMultiple(driver)
        # start WhsBkg checking process

        if builtins.SOCount > 0:
            # initialized DF DESt, DR, DS
            if builtins.SOCount == 1:
                builtins.gSOGrpType = "Single"
                if isinstance(builtins.VendorBkg, str):
                    UpdateSinglSOtoDF(builtins.VendorBkg)
                else:
                    for k, so in enumerate(builtins.VendorBkg.itertuples(), 1):
                        newso = so[1]
                        UpdateSinglSOtoDF(newso)
            if builtins.SOCount > 1:
                builtins.gSOGrpType = "Multiple"
            if builtins.gSOGrpType == "Multiple":
                for k, so in enumerate(builtins.VendorBkg.itertuples(), 1):
                    newso = so[1]
                    OpenFindWhsBkg(driver)
                    time.sleep(5)
                    WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, '#quickSearchText'))).click()
                    klen = len(builtins.VendorBkg)

                    soIndx = 0
                    ManageWhsBkg(newso, driver)
                    time.sleep(3)

                UniqueBL = GetUniqueVals(builtins.dataframe['BL'])
                # checking of upload type depending on values of Doc Receipt & Doc Scan Date
                CheckUploadType()
                # consolidate all related files
                remarks = ConsolidateFiles(gCurrBL)
                txtConsolidated = "Consolidated"
                if txtConsolidated in remarks:
                    # update remarks on DF
                    UpdateToolRemarks(gCurrBL, remarks)
                    # zip to be uploaded files

                    ZipTheDocs(gCurrBL)
                    # process steps for External Docs process
                    OpenExternalDocs(driver)
                    ManageExternalDocs(gCurrBL, driver)

                else:
                    # update remarks on DF
                    UpdateToolRemarks(gCurrBL, remarks)
                    # # zip to be uploaded files
                    #
                    # ZipTheDocs(gCurrBL)
                    # # process steps for External Docs process
                    # OpenExternalDocs(driver)
                    # ManageExternalDocs(gCurrBL, driver)
            else:
                OpenFindWhsBkg(driver)
                time.sleep(3)
                if IsPageReady(driver):
                    klen = len(builtins.VendorBkg)

                soIndx = 0
                if isinstance(builtins.VendorBkg, str):
                    newso = builtins.VendorBkg
                else:
                    for k, so in enumerate(builtins.VendorBkg.itertuples(), 1):
                        newso = so[1]
                ManageWhsBkg(newso, driver)
                time.sleep(3)

                UniqueBL = GetUniqueVals(builtins.dataframe['BL'])
                # checking of upload type depending on values of Doc Receipt & Doc Scan Date
                CheckUploadType()
                # consolidate all related files
                remarks = ConsolidateFiles(gCurrBL)
                # if remarks.find("not found") > 0:
                #     UpdateToolRemarks(gCurrBL, remarks)
                #
                # else:
                # update remarks on DF
                UpdateToolRemarks(gCurrBL, remarks)

                # zip to be uploaded files

                ZipTheDocs(gCurrBL)
                # process steps for External Docs process
                OpenExternalDocs(driver)
                ManageExternalDocs(gCurrBL, driver)
        else:
            gRemarks = 'BL Not found'
            UpdateToolRemarks(gCurrBL, gRemarks)

        # post upload process here
    ManagePostProcessing('.')
    driver.quit